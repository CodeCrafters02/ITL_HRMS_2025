from rest_framework import viewsets, generics, status
from rest_framework.exceptions import ValidationError
from django.shortcuts import get_object_or_404
import pytz
import string
from rest_framework.decorators import action
from rest_framework.response import Response
from django.utils.timezone import localtime
from django.db.models import Sum, Q, Count
from django.utils import timezone
from datetime import datetime, timedelta, time
from calendar import monthrange
from django.db import transaction
import calendar
from django.contrib.auth import authenticate
from rest_framework_simplejwt.tokens import RefreshToken
from datetime import date
import io
from django.utils import timezone
from .utils import generate_letter_pdf, fill_placeholders, validate_geofence, compute_attendance_metrics, generate_payslip_pdf, compute_loan_emi, compute_asset_deductions, compute_loan_disbursement, compute_salary_structure_allowances
from decimal import Decimal, ROUND_HALF_UP
from django.core.mail import EmailMessage
from django.core.files.base import ContentFile
import uuid
from rest_framework.views import APIView
from rest_framework.response import Response
import re
from rest_framework.permissions import IsAuthenticated
from .permissions import IsMaster, IsAdminUser, IsCompanyChatUser, CanReadCompanyCalendar
from .serializers import *
from .models import *

from rest_framework import filters
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django_filters.rest_framework import DjangoFilterBackend
from .chat_serializers import (
    ChatConversationSerializer,
    ChatMessageSerializer,
)


class CustomPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100
class MyIPAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return Response({'ip': ip})


class CustomPasswordChangeAPIView(generics.UpdateAPIView):
    serializer_class = CustomPasswordChangeSerializer
    permission_classes = [IsAuthenticated]

    def update(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({"detail": "Password updated successfully."}, status=status.HTTP_200_OK)


class MasterRegisterViewSet(viewsets.ModelViewSet):
    queryset = UserRegister.objects.filter(role='master')
    serializer_class = UserRegisterSerializer
    permission_classes = [permissions.AllowAny]

    def create(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        email = serializer.validated_data.get('email')
        if UserRegister.objects.filter(email=email).exists():
            return Response({"detail": "Email already exists."}, status=status.HTTP_400_BAD_REQUEST)
        if serializer.validated_data['role'] != 'master':
            return Response({"error": "Role must be 'master'"}, status=status.HTTP_400_BAD_REQUEST)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
class UserManagementViewSet(viewsets.ModelViewSet):
    serializer_class = UserSerializer
    # permission_classes = [permissions.IsAuthenticated, IsMaster]
    permission_classes = [permissions.AllowAny]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['username', 'email', 'role']

    def get_queryset(self):
        queryset = UserRegister.objects.all().order_by('id')
        created_by = self.request.query_params.get('created_by')
        if created_by:
            queryset = queryset.filter(created_by=created_by)
        return queryset

    def perform_create(self, serializer):
        user = self.request.user
        if user.is_authenticated:
            serializer.save(created_by=user)
        else:
            serializer.save()
class AdminRegisterViewSet(viewsets.ModelViewSet):
    serializer_class = AdminRegisterSerializer
    permission_classes = [IsAuthenticated, IsMaster]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['username', 'email']

    def get_queryset(self):
        unassigned = self.request.query_params.get('unassigned') == 'true'
        include_id = self.request.query_params.get('include_id')
        
        queryset_filter = Q(role='admin')
        
        if unassigned:
            if include_id:
                queryset_filter &= (Q(company__isnull=True) | Q(id=include_id))
            else:
                queryset_filter &= Q(company__isnull=True)
            
        return UserRegister.objects.filter(queryset_filter).order_by('id')

    def update(self, request, pk=None, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        try:
            admin = UserRegister.objects.get(pk=pk, role='admin')
        except UserRegister.DoesNotExist:
            return Response({'detail': 'Admin not found.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = AdminRegisterSerializer(admin, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)

    def create(self, request):
        serializer = AdminRegisterSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
class PasswordChangeView(generics.UpdateAPIView):
    serializer_class = PasswordChangeSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        return self.request.user

    def update(self, request, *args, **kwargs):
        user = self.get_object()
        serializer = self.get_serializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)

        user.set_password(serializer.validated_data['new_password'])
        user.save()

        return Response({"detail": "Password updated successfully."}, status=status.HTTP_200_OK)

class MasterDashboardView(APIView):
    permission_classes = [IsAuthenticated,IsMaster]

    def get(self, request):
        user = request.user

        if user.role != 'master':
            return Response(
                {"detail": "You are not authorized for this dashboard."},
                status=status.HTTP_403_FORBIDDEN
            )

        companies_data = []
        companies = Company.objects.all()

        for company in companies:
            
            admins = UserRegister.objects.filter(company=company, role='admin')

            admin_serializer = MasterDashboardSerializer(admins, many=True)
            logo_url = request.build_absolute_uri(company.logo.url) if company.logo else None


            companies_data.append({
                "id": company.id,
                "name": company.name,
                "address": company.address,
                "location": company.location,
                "email": company.email,
                "phone_number": company.phone_number,
                "logo": logo_url,
                "admins": admin_serializer.data,
            })

        return Response({
            "companies": companies_data,
            "total_companies": companies.count(),
            "total_admins": UserRegister.objects.filter(role='admin').count(),
            "total_masters": UserRegister.objects.filter(role='master').count(),
            "total_employees": UserRegister.objects.filter(role='employee').count()
        })

def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

class LoginAPIView(APIView):

    def post(self, request):
        username = request.data.get("username")
        password = request.data.get("password")

        user = None
        try:
            # allow login with either username or email
            user_obj = UserRegister.objects.get(Q(username=username) | Q(email=username))
            user = authenticate(username=user_obj.username, password=password)
        except UserRegister.DoesNotExist:
            user = None
        except UserRegister.MultipleObjectsReturned:
            # fallback if multiple matching emails exists
            user_obj = UserRegister.objects.filter(Q(username=username) | Q(email=username)).first()
            user = authenticate(username=user_obj.username, password=password)

        if user is not None:
            if not user.is_active:
                return Response({"detail": "User account is disabled."}, status=status.HTTP_403_FORBIDDEN)

            # ✅ Issue JWT tokens
            refresh = RefreshToken.for_user(user)
            access_token = str(refresh.access_token)
            refresh_token = str(refresh)

            return Response({
                "access": access_token,
                "refresh": refresh_token,
                "id": user.id,
                "username": user.username,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "role": user.role,
                "is_reporting_manager": user.is_reporting_manager
            }, status=status.HTTP_200_OK)

        return Response({"detail": "Invalid credentials"}, status=status.HTTP_401_UNAUTHORIZED)
     
from rest_framework import filters
from rest_framework.pagination import PageNumberPagination

class CustomPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class CompanyWithAdminViewSet(viewsets.ModelViewSet):
    queryset = Company.objects.all().order_by('id')
    serializer_class = CompanyWithAdminSerializer
    permission_classes = [IsAuthenticated, IsMaster]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'email', 'location', 'phone_number']

    def get_serializer_context(self):
        return {'request': self.request}


class CompanyLogoAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        company = getattr(request.user, 'company', None)
        if not company:
            return Response({'detail': 'No company found.'}, status=404)
        
        serializer = CompanyWithAdminSerializer(company, context={'request': request})
        return Response(serializer.data)


class CompanyUpdateAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]
    
    def get(self, request):
       
        user = request.user
        company = getattr(user, 'company', None)
        if not company:
            return Response({"detail": "No company found for user."}, status=status.HTTP_404_NOT_FOUND)
        serializer = CompanySerializer(company, context={"request": request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    def put(self, request, pk):
        try:
            company = Company.objects.get(pk=pk)
        except Company.DoesNotExist:
            return Response({"detail": "Company not found."}, status=status.HTTP_404_NOT_FOUND)

        
        serializer = CompanySerializer(company, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class AdminDashboardAPIView(APIView):
    permission_classes = [IsAuthenticated,IsAdminUser]
    def get(self, request):
        tz = pytz.timezone('Asia/Kolkata')
        now = timezone.localtime(timezone.now(), tz)
        today = now.date()
        company = request.user.company

        # Personal birthday message for the admin
        birthday_message = None
        emp_profile = request.user.employee_profile
        if emp_profile and emp_profile.date_of_birth:
            if emp_profile.date_of_birth.day == today.day and emp_profile.date_of_birth.month == today.month:
                birthday_message = f"Happy Birthday, {request.user.first_name or request.user.username}! 🎉"

        # Department count
        total_departments = Department.objects.filter(company=company).count()

        # Leaves today
        leaves_today = EmpLeave.objects.filter(
            company=company,
            from_date__lte=today,
            to_date__gte=today,
            status='Approved'
        ).count()

        # Employee Overview (excluding admin/master roles)
        base_employee_qs = Employee.objects.filter(company=company, is_active=True).exclude(user__role__in=['admin', 'master'])
        total_employees = base_employee_qs.count()
        active_employees_count = total_employees
        inactive_employees = Employee.objects.filter(company=company, is_active=False).exclude(user__role__in=['admin', 'master']).count()
        new_joinees = Employee.objects.filter(
            company=company,
            date_of_joining__year=today.year,
            date_of_joining__month=today.month
        ).exclude(user__role__in=['admin', 'master']).count()
        
        exits_this_month = Employee.objects.filter(
            company=company,
            relieved_info__relieving_date__year=today.year,
            relieved_info__relieving_date__month=today.month
        ).exclude(user__role__in=['admin', 'master']).count()
        # Upcoming Birthdays (next 30 days)
        # NOTE: month/day range filtering breaks across month/year boundaries.
        # Compute each employee's next birthday date and filter in Python (company sizes are typically manageable).
        next_30 = today + timezone.timedelta(days=30)
        upcoming_birthdays = []
        for e in base_employee_qs.only('id', 'first_name', 'last_name', 'date_of_birth'):
            dob = e.date_of_birth
            if not dob:
                continue
            try:
                next_bday = dob.replace(year=today.year)
            except ValueError:
                # Handle Feb 29 in non-leap years: treat as Feb 28
                next_bday = timezone.datetime(today.year, 2, 28).date()
            if next_bday < today:
                try:
                    next_bday = next_bday.replace(year=today.year + 1)
                except ValueError:
                    next_bday = timezone.datetime(today.year + 1, 2, 28).date()
            if today <= next_bday <= next_30:
                upcoming_birthdays.append((next_bday, e))
        upcoming_birthdays.sort(key=lambda x: x[0])
        

        # Attendance Snapshot - Calculate based on shift timing
        employees_for_attendance = base_employee_qs
        attendance_snapshot = self._calculate_attendance_snapshot(company, today, employees_for_attendance)

        # Pending Leave Requests
        pending_leaves = EmpLeave.objects.filter(company=company, status='Pending').count()

        # Payroll Status
        current_month = today.month
        current_year = today.year
        payroll_batches = PayrollBatch.objects.filter(company=company, month=current_month, year=current_year)
        payroll_status = "pending"
        if payroll_batches.filter(status='Locked').exists():
            payroll_status = "completed"

        # Upcoming Salary Release (next batch with status 'Draft')
        next_salary_release = payroll_batches.filter(status='Draft').order_by('id').first()
        next_salary_release_date = None
        if next_salary_release:
            next_salary_release_date = f"{current_year}-{current_month}-01"  # Or use a real field if you have one

        return Response({
            "department_count": total_departments,
            "leaves_today": leaves_today,
            "employee_overview": {
                "total": total_employees,
                "active": active_employees_count,
                "inactive": inactive_employees,
                "new_joinees": new_joinees,
                "exits_this_month": exits_this_month,
            },
            "upcoming_birthdays": [
                {
                    "name": f"{(e.first_name or '').strip()} {(e.last_name or '').strip()}".strip() or str(e.employee_id or e.id),
                    "date_of_birth": e.date_of_birth,
                }
                for _, e in upcoming_birthdays
            ],
           
            "attendance_snapshot": attendance_snapshot,
            "pending_leave_requests": pending_leaves,
            "payroll_status": payroll_status,
            "next_salary_release_date": next_salary_release_date,
            "birthday_message": birthday_message,
        })

    def _calculate_attendance_snapshot(self, company, today, employees):
        """
        Calculate attendance snapshot based on shift timing and current time.
        Returns counts for present, absent, half_day, full_day_leave, on_leave
        """
        present = 0
        absent = 0
        half_day = 0
        full_day_leave = 0
        on_leave = 0
        
        # Get current time in local timezone
        now = timezone.localtime(timezone.now())
        current_time = now.time()
        
        # Get default grace period from company's shift policies
        default_grace_period = self._get_company_default_grace_period(company)
        
        # Get approved leaves for today
        approved_leaves = EmpLeave.objects.filter(
            company=company,
            from_date__lte=today,
            to_date__gte=today,
            status='Approved'
        ).values_list('employee_id', flat=True)
        
        for employee in employees:
            # Check if employee is on approved leave
            if employee.id in approved_leaves:
                on_leave += 1
                continue
                
            # Get employee's attendance record for today
            attendance = Attendance.objects.filter(
                employee=employee, 
                date=today
            ).first()
            
            # Get employee's shift
            shift = getattr(employee, 'shift_assigned', None)
            
            if attendance and attendance.check_in:
                # Employee has checked in - determine status based on check-in time
                check_in_local = timezone.localtime(attendance.check_in)
                check_in_time = check_in_local.time()
                
                if shift:
                    shift_start = shift.checkin
                    grace_period = shift.grace_period or default_grace_period
                    half_day_duration = shift.half_day or timedelta(hours=4)
                    
                    # Calculate half day time (shift start + half day duration)
                    half_day_time = (datetime.combine(today, shift_start) + half_day_duration).time()
                    grace_time = (datetime.combine(today, shift_start) + grace_period).time()
                    
                    if check_in_time >= half_day_time:
                        # Checked in at or after half day time
                        full_day_leave += 1
                    elif check_in_time >= grace_time:
                        # Checked in at or after grace period but before half day
                        half_day += 1
                    else:
                        # Checked in within grace period
                        present += 1
                else:
                    # No shift assigned - use default logic with company grace period
                    default_shift_start = time(9, 0)  # Assume 9 AM default start
                    default_half_day = timedelta(hours=4)  # 4 hours for half day
                    half_day_time = (datetime.combine(today, default_shift_start) + default_half_day).time()
                    grace_time = (datetime.combine(today, default_shift_start) + default_grace_period).time()
                    
                    if check_in_time >= half_day_time:
                        full_day_leave += 1
                    elif check_in_time >= grace_time:
                        half_day += 1
                    else:
                        present += 1
            else:
                # Employee hasn't checked in - determine status based on current time
                if shift:
                    shift_start = shift.checkin
                    grace_period = shift.grace_period or default_grace_period
                    half_day_duration = shift.half_day or timedelta(hours=4)
                    
                    grace_time = (datetime.combine(today, shift_start) + grace_period).time()
                    half_day_time = (datetime.combine(today, shift_start) + half_day_duration).time()
                    
                    if current_time >= half_day_time:
                        # Past half day time and no check-in
                        full_day_leave += 1
                    elif current_time >= grace_time:
                        # Past grace period but before half day
                        half_day += 1
                    else:
                        # Still within grace period
                        absent += 1
                else:
                    # No shift assigned - use default logic with company grace period
                    default_shift_start = time(9, 0)  # Assume 9 AM default start
                    default_half_day = timedelta(hours=4)  # 4 hours for half day
                    grace_time = (datetime.combine(today, default_shift_start) + default_grace_period).time()
                    half_day_time = (datetime.combine(today, default_shift_start) + default_half_day).time()
                    
                    if current_time >= half_day_time:
                        full_day_leave += 1
                    elif current_time >= grace_time:
                        half_day += 1
                    else:
                        absent += 1
        
        return {
            "present": present,
            "absent": absent,
            "half_day": half_day,
            "full_day_leave": full_day_leave,
            "on_leave": on_leave,
        }

    def _get_company_default_grace_period(self, company):
        """
        Get the default grace period for the company from its shift policies.
        Returns the minimum grace period found, or 15 minutes as fallback.
        """
        from .models import ShiftPolicy
        
        # Get all shift policies for the company
        shift_policies = ShiftPolicy.objects.filter(company=company)
        
        if shift_policies.exists():
            # Find the minimum grace period among all shift policies
            grace_periods = []
            for shift in shift_policies:
                if shift.grace_period:
                    grace_periods.append(shift.grace_period)
            
            if grace_periods:
                # Return the minimum grace period
                return min(grace_periods)
            else:
                # No grace periods set, use default
                return timedelta(minutes=15)
        else:
            # No shift policies exist, use default
            return timedelta(minutes=15)

class DepartmentViewSet(viewsets.ModelViewSet):
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated, IsAdminUser | IsMaster]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['department_name']
    
    def get_queryset(self):
        user = self.request.user
        if user.role == 'master':
            return Department.objects.all().order_by('id')
        return Department.objects.filter(company=user.company).order_by('id')

    def perform_create(self, serializer):
        company = self.request.user.company
        if not company:
            raise serializers.ValidationError({"company": "No company found for the current user."})
        serializer.save(company=company)


class LevelViewSet(viewsets.ModelViewSet):
    serializer_class = LevelSerializer
    permission_classes = [IsAuthenticated, IsAdminUser | IsMaster]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['level_name', 'description']
   
    def get_queryset(self):
        user = self.request.user
        if user.role == 'master':
            return Level.objects.all().order_by('id')
        return Level.objects.filter(company=user.company).order_by('id')

    def perform_create(self, serializer):
        company = self.request.user.company
        if not company:
            raise serializers.ValidationError({"company": "No company found for the current user."})
        serializer.save(company=company)


class DesignationViewSet(viewsets.ModelViewSet):
    serializer_class = DesignationSerializer
    permission_classes = [IsAuthenticated, IsAdminUser | IsMaster]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['designation_name', 'department__department_name', 'level__level_name']
   
    def get_queryset(self):
        user = self.request.user
        if user.role == 'master':
            return Designation.objects.select_related('department', 'level').order_by('id')
        return Designation.objects.filter(company=user.company).select_related('department', 'level').order_by('id')

    def perform_create(self, serializer):
        company = self.request.user.company
        if not company:
            raise serializers.ValidationError({"company": "No company found for the current user."})
        serializer.save(company=company)        

class DesignationSalaryViewSet(viewsets.ModelViewSet):
    serializer_class = DesignationSalarySerializer
    permission_classes = [IsAuthenticated, IsAdminUser | IsMaster]
    
    def get_queryset(self):
        user = self.request.user
        qs = DesignationSalary.objects.all().select_related('designation', 'designation__department')
        if user.role != 'master':
            qs = qs.filter(company=user.company)
        
        dept_id = self.request.query_params.get('department')
        if dept_id:
            qs = qs.filter(designation__department_id=dept_id)
            
        return qs

    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company)

class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = [
        'employee_id',
        'first_name',
        'middle_name',
        'last_name',
        'email',
        'mobile',
        'department__department_name',
        'designation__designation_name',
        'level__level_name',
    ]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def toggle_work_location(self, request, pk=None):
        if not (request.user.role == 'admin' or request.user.role == 'master'):
            return Response({"detail": "Only admins can manually toggle work location."}, status=status.HTTP_403_FORBIDDEN)
        
        employee = self.get_object()
        old_loc = employee.work_location
        new_loc = 'home' if old_loc == 'office' else 'office'
        reason = request.data.get('reason', f"Direct change by {request.user.username}")

        employee.work_location = new_loc
        employee.save()

        # Create log
        WorkLocationLog.objects.create(
            employee=employee,
            from_location=old_loc,
            to_location=new_loc,
            changed_by=request.user.employee_profile,
            reason=reason
        )

        return Response({
            "detail": f"Successfully changed {employee.full_name}'s location to {new_loc}",
            "work_location": new_loc
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def me(self, request):
        emp_profile = getattr(request.user, 'employee_profile', None)
        if not emp_profile:
            return Response({"error": "No employee profile found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = self.get_serializer(emp_profile)
        return Response(serializer.data)

    def _ensure_employee_profiles_for_company(self, company):
        """
        Ensure every employee `UserRegister` in this company has an `Employee` profile.
        This fixes mismatches where SSO-created users exist without an Employee record,
        causing Admin "Employee Register" to appear incomplete.
        """
        if not company:
            return

        # Users with role=employee but no Employee profile (OneToOne reverse is `employee` by default)
        missing_profile_users = (
            UserRegister.objects.filter(role__in=["employee", "admin"], company=company, employee__isnull=True)
            .only("id", "first_name", "last_name", "email", "company_id")
        )

        for u in missing_profile_users:
            if not u.email:
                continue

            # Prefer linking an existing Employee row (created earlier via HR) by email+company
            existing = (
                Employee.objects.filter(company=company, email__iexact=u.email)
                .select_related("company")
                .first()
            )
            if existing:
                if not existing.user_id:
                    existing.user_id = u.id
                    # Backfill basic fields if empty
                    if not existing.first_name:
                        existing.first_name = u.first_name or existing.first_name
                    if not existing.last_name:
                        existing.last_name = u.last_name or existing.last_name
                    if not existing.email:
                        existing.email = u.email
                    existing.save(update_fields=["user", "first_name", "last_name", "email"])
                continue

            # Otherwise create a lightweight Employee profile
            Employee.objects.create(
                user_id=u.id,
                company=company,
                first_name=u.first_name or None,
                last_name=u.last_name or None,
                email=u.email,
            )

    def get_permissions(self):
        # Employees can view (list/retrieve) employees from their company,
        # but cannot create/update/delete.
        if self.action in ['list', 'retrieve', 'get_reporting_manager_choices', 'me']:
            permission_classes = [IsAuthenticated]
        else:
            permission_classes = [IsAuthenticated, IsAdminUser | IsMaster]
        return [permission() for permission in permission_classes]

    def get_queryset(self):
        user = self.request.user
        qs = Employee.objects.all()
        if user.role != 'master':
            company = getattr(user, 'company', None)
            if not company:
                # Fallback (SSO/legacy): infer company from linked employee profile
                emp = Employee.objects.filter(user=user).select_related('company').first()
                company = emp.company if emp else None
            if company:
                # Keep the Employee register complete for this company
                self._ensure_employee_profiles_for_company(company)
                qs = qs.filter(company=company)
                include_admins = self.request.query_params.get('include_admins', 'false').lower() == 'true'
                if not include_admins:
                    qs = qs.exclude(user__role__in=['admin', 'master'])
            else:
                qs = qs.none()
            
        return qs.select_related(
            'department', 'designation', 'level', 'reporting_manager', 
            'reporting_level', 'shift_assigned'
        ).order_by('employee_id')

    def perform_destroy(self, instance):
        """
        When deleting an Employee, also delete the linked UserRegister account.
        The OneToOneField CASCADE only works UserRegister→Employee (not reverse),
        so we must manually clean up the UserRegister to avoid orphaned login accounts.
        """
        linked_user = instance.user
        instance.delete()
        if linked_user:
            linked_user.delete()

    @action(detail=False, methods=['get'], url_path='get-reporting-manager-choices')
    def get_reporting_manager_choices(self, request):
        company = getattr(request.user, 'company', None)
        if not company:
            return Response({"error": "User has no company"}, status=status.HTTP_400_BAD_REQUEST)

        reporting_level_id = request.query_params.get('reporting_level_id')

        # Fetch all levels for this company
        levels = Level.objects.filter(company=company).order_by('id')
        if reporting_level_id:
            levels = levels.filter(id=reporting_level_id)

        response_data = []

        for level in levels:
            # Find all designations for this level
            designations = Designation.objects.filter(company=company, level=level)

            # Get all employees belonging to these designations
            employees = Employee.objects.filter(
                company=company,
                designation__in=designations
            ).order_by('id')

            level_data = {
                "level_id": level.id,
                "level_name": level.level_name,
                "employees": [
                    {
                        "id": emp.id,
                        "name": emp.full_name or emp.user.first_name or emp.user.username
                    }
                    for emp in employees
                ]
            }

            response_data.append(level_data)

        return Response(response_data, status=status.HTTP_200_OK)

    
class SupplyItemViewSet(viewsets.ModelViewSet):
    serializer_class = SupplyItemSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    pagination_class = CustomPagination

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsAdminUser()]
    filter_backends = [filters.SearchFilter]
    search_fields = ['item_code', 'item_name', 'vendor_details', 'sub_category']

    def get_queryset(self):
        return SupplyItem.objects.filter(company=self.request.user.company).order_by('item_code')


class FixedAssetViewSet(viewsets.ModelViewSet):
    serializer_class = FixedAssetSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    filter_backends = [filters.SearchFilter]
    search_fields = ['asset_tag', 'serial_number', 'model_brand', 'category', 'status']

    def get_queryset(self):
        qs = FixedAsset.objects.filter(company=self.request.user.company).select_related(
            'assigned_to', 'variable_supply_item'
        )
        if getattr(self.request.user, 'role', None) == 'employee':
            emp = getattr(self.request.user, 'employee_profile', None)
            if emp:
                qs = qs.filter(assigned_to=emp)
            else:
                qs = qs.none()
        return qs

    def perform_create(self, serializer):
        if self.request.user.role != 'admin':
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Only admins can create assets.")
        serializer.save()

    def perform_update(self, serializer):
        if self.request.user.role != 'admin':
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Only admins can update assets.")
        print(f"DEBUG: FixedAsset update data: {serializer.validated_data}")
        serializer.save()

    def perform_destroy(self, instance):
        if self.request.user.role != 'admin':
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Only admins can delete assets.")
        instance.delete()


class AssetRequestViewSet(viewsets.ModelViewSet):
    serializer_class = AssetRequestSerializer
    pagination_class = CustomPagination
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    filterset_fields = ['approval_status', 'batch_id']
    search_fields = ['remarks', 'approval_status', 'requested_by__first_name', 'requested_by__last_name']

    def get_permissions(self):
        # Employees can create requests and view their own.
        if self.action in ("create", "list", "retrieve"):
            return [IsAuthenticated()]
        # Only admins can approve/edit/delete requests
        return [IsAuthenticated(), IsAdminUser()]

    def get_queryset(self):
        qs = AssetRequest.objects.filter(company=self.request.user.company).select_related(
            'requested_by', 'related_fixed_asset', 'related_supply_item'
        )
        if getattr(self.request.user, 'role', None) == 'employee':
            emp = getattr(self.request.user, 'employee_profile', None)
            if emp:
                qs = qs.filter(requested_by=emp)
            else:
                qs = qs.none()
        return qs
    def perform_create(self, serializer):
        u = self.request.user
        emp = getattr(u, 'employee_profile', None) if getattr(u, 'role', None) == 'employee' else None
        
        supply_item = serializer.validated_data.get('related_supply_item')
        qty = serializer.validated_data.get('requested_quantity', 1)
        
        if supply_item:
            if supply_item.available_quantity < qty:
                from rest_framework.exceptions import ValidationError
                raise ValidationError({"detail": f"Out of stock. Only {supply_item.available_quantity} available."})
            if supply_item.max_per_order and qty > supply_item.max_per_order:
                from rest_framework.exceptions import ValidationError
                raise ValidationError({"detail": f"Maximum allowed is {supply_item.max_per_order} per order."})

            supply_item.available_quantity -= qty
            supply_item.save()
            
        if emp:
            serializer.save(company=u.company, requested_by=emp)
        else:
            serializer.save(company=u.company)

    @action(detail=False, methods=['post'], url_path='batch-action')
    def batch_action(self, request):
        batch_id = request.data.get('batch_id')
        new_status = request.data.get('status')
        if not batch_id or not new_status:
            return Response({"detail": "batch_id and status required."}, status=400)
            
        from django.db.models import Q
        reqs = AssetRequest.objects.filter(
            Q(batch_id=batch_id) | Q(id=batch_id if str(batch_id).isdigit() else -1),
            company=request.user.company
        )
        if not reqs.exists():
            return Response({"detail": f"No requests found for ID/Batch: {batch_id}."}, status=404)

        for r in reqs:
            # If changing from pending to rejected, add stock back
            if new_status == 'rejected' and r.approval_status != 'rejected':
                if r.related_supply_item:
                    r.related_supply_item.available_quantity += r.requested_quantity
                    r.related_supply_item.save()
            
            # If changing from rejected back to approved/pending, reduce stock (if enough)
            elif new_status != 'rejected' and r.approval_status == 'rejected':
                if r.related_supply_item:
                    if r.related_supply_item.available_quantity < r.requested_quantity:
                         pass
                    r.related_supply_item.available_quantity -= r.requested_quantity
                    r.related_supply_item.save()

            r.approval_status = new_status
            if new_status == 'approved':
                # Capture action type and payment amount for any asset type if provided
                r.admin_action_type = request.data.get('admin_action_type', r.admin_action_type)
                payment_val = request.data.get('employee_payment_amount')
                if payment_val is not None:
                    try:
                        r.employee_payment_amount = Decimal(str(payment_val))
                    except:
                        pass
            r.save()

        return Response({"detail": f"Batch {batch_id} set to {new_status}."})

    def perform_update(self, serializer):
        instance = self.get_object()
        new_status = serializer.validated_data.get('approval_status')
        old_status = instance.approval_status
        # Capture asset details if approving
        if new_status == 'approved':
            serializer.validated_data['admin_action_type'] = self.request.data.get('admin_action_type', instance.admin_action_type)
            payment_val = self.request.data.get('employee_payment_amount')
            if payment_val is not None:
                try:
                    serializer.validated_data['employee_payment_amount'] = Decimal(str(payment_val))
                except:
                    pass
        
        serializer.save()
       # If changing TO rejected from something else -> add stock back
        if new_status == 'rejected' and old_status != 'rejected':
            if instance.related_supply_item:
                instance.related_supply_item.available_quantity += instance.requested_quantity
                instance.related_supply_item.save()
        
        # If changing FROM rejected to something else -> reduce stock
        elif new_status != 'rejected' and old_status == 'rejected':
            if instance.related_supply_item:
                instance.related_supply_item.available_quantity -= instance.requested_quantity
                instance.related_supply_item.save()
                
        serializer.save()

    def perform_destroy(self, instance):
        # If we delete a request that wasn't rejected, return the reserved stock
        if instance.approval_status != 'rejected' and instance.related_supply_item:
            instance.related_supply_item.available_quantity += instance.requested_quantity
            instance.related_supply_item.save()
        instance.delete()

class AssetSupportingDocumentViewSet(viewsets.ModelViewSet):
    serializer_class = AssetSupportingDocumentSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    pagination_class = CustomPagination
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def get_queryset(self):
        qs = AssetSupportingDocument.objects.filter(company=self.request.user.company)
        fa = self.request.query_params.get('fixed_asset')
        si = self.request.query_params.get('supply_item')
        ar = self.request.query_params.get('asset_request')
        if fa:
            qs = qs.filter(fixed_asset_id=fa)
        if si:
            qs = qs.filter(supply_item_id=si)
        if ar:
            qs = qs.filter(asset_request_id=ar)
        return qs.order_by('-uploaded_at')

    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company, uploaded_by=self.request.user)

class ReimbursementCategoryViewSet(viewsets.ModelViewSet):
    queryset = ReimbursementCategory.objects.all()
    serializer_class = ReimbursementCategorySerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'description']
    
    def get_queryset(self):
        user = self.request.user
        if user.role == 'master':
            return ReimbursementCategory.objects.all().order_by('id')
        return ReimbursementCategory.objects.filter(company=user.company).order_by('id')

    def perform_create(self, serializer):
        user = self.request.user
        if user.role == 'master':
            serializer.save()
        else:
            serializer.save(company=user.company)

class ReimbursementRequestViewSet(viewsets.ModelViewSet):
    serializer_class = ReimbursementRequestSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['employee__first_name', 'employee__last_name', 'category__name', 'custom_category', 'description', 'status']

    def get_queryset(self):
        user = self.request.user
        company = user.company
        emp = getattr(user, 'employee_profile', None)
        
        if user.role == 'admin' or user.role == 'master':
            queryset = ReimbursementRequest.objects.filter(company=company)
        elif emp:
            queryset = ReimbursementRequest.objects.filter(
                Q(employee=emp) | Q(reporting_manager=emp)
            )
        else:
            queryset = ReimbursementRequest.objects.none()
            
        # Advanced Filtering
        status_param = self.request.query_params.get('status')
        if status_param:
            queryset = queryset.filter(status=status_param)

        employee_param = self.request.query_params.get('employee')
        if employee_param:
            queryset = queryset.filter(employee_id=employee_param)

        manager_param = self.request.query_params.get('reporting_manager')
        if manager_param:
            queryset = queryset.filter(reporting_manager_id=manager_param)

        start_date = self.request.query_params.get('start_date')
        if start_date:
            queryset = queryset.filter(created_at__date__gte=start_date)

        end_date = self.request.query_params.get('end_date')
        if end_date:
            queryset = queryset.filter(created_at__date__lte=end_date)
            
        return queryset.order_by('-created_at')

    @action(detail=False, methods=['get'])
    def stats(self, request):
        user = self.request.user
        if user.role not in ['admin', 'master']:
            return Response({"detail": "Not authorized."}, status=status.HTTP_403_FORBIDDEN)
            
        queryset = self.get_queryset()
        
        # Overall totals
        overall = queryset.aggregate(
            total_amount=Sum('amount'),
            approved_amount=Sum('amount', filter=Q(status='approved')),
            pending_amount=Sum('amount', filter=Q(status='pending'))
        )

        # Employee-wise totals (only for approved ones usually, or based on filters)
        employee_wise = queryset.values(
            'employee_id', 
            'employee__first_name', 
            'employee__last_name'
        ).annotate(
            total=Sum('amount'),
            approved=Sum('amount', filter=Q(status='approved')),
            count=Count('id')
        ).order_by('-total')

        return Response({
            "overall": overall,
            "employee_wise": employee_wise
        })

    def perform_create(self, serializer):
        emp = self.request.user.employee_profile
        if not emp:
            raise serializers.ValidationError("Employee profile required to request reimbursement.")
        
        category_id = self.request.data.get('category')
        if category_id:
            try:
                category = ReimbursementCategory.objects.get(id=category_id)
                if category.min_tenure_months > 0:
                    if not emp.date_of_joining:
                        raise serializers.ValidationError(f"This category requires {category.min_tenure_months} months of tenure. Your joining date is not set.")
                    
                    # Calculate tenure in months
                    today = timezone.localdate()
                    joining = emp.date_of_joining
                    tenure_months = (today.year - joining.year) * 12 + (today.month - joining.month)
                    
                    # Adjust if day of month hasn't reached yet
                    if today.day < joining.day:
                        tenure_months -= 1
                        
                    if tenure_months < category.min_tenure_months:
                        raise serializers.ValidationError(f"You must complete at least {category.min_tenure_months} months of tenure to apply for this reimbursement. Current tenure: {tenure_months} months.")
            except ReimbursementCategory.DoesNotExist:
                pass

        serializer.save(
            employee=emp,
            company=emp.company,
            reporting_manager=emp.reporting_manager,
            status='pending'
        )

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        reimbursement = self.get_object()
        user = request.user
        emp = getattr(user, 'employee_profile', None)
        
        if user.role != 'admin' and (not emp or reimbursement.reporting_manager != emp):
            return Response({"detail": "Not authorized to approve this request."}, status=status.HTTP_403_FORBIDDEN)
            
        reimbursement.status = 'approved'
        reimbursement.save()
        return Response({"status": "approved"})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        reimbursement = self.get_object()
        user = request.user
        emp = getattr(user, 'employee_profile', None)
        reason = request.data.get('rejection_reason', '')

        if user.role != 'admin' and (not emp or reimbursement.reporting_manager != emp):
            return Response({"detail": "Not authorized to reject this request."}, status=status.HTTP_403_FORBIDDEN)

        reimbursement.status = 'rejected'
        reimbursement.rejection_reason = reason
        reimbursement.save()
        return Response({"status": "rejected"})

    
class RecruitmentViewSet(viewsets.ModelViewSet):
    queryset = Recruitment.objects.all()
    serializer_class = RecruitmentSerializer
    permission_classes = [IsAuthenticated, IsAdminUser | IsMaster]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['reference_id', 'name', 'email', 'job_title', 'status', 'guardian_name']

    def perform_update(self, serializer):
        instance = serializer.save()

        # Trigger email only if status field is updated
        new_status = self.request.data.get('status')
        if new_status:
            if new_status == 'rejected':
                send_mail(
                    subject='Application Status',
                    message=(
                        f"Dear {instance.name},\n\n"
                        f"We regret to inform you that you were not selected for the position of {instance.job_title}.\n"
                        "We wish you all the best in your future endeavors.\n\n"
                        "Regards,\nYour Company"
                    ),
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[instance.email],
                    fail_silently=False,
                )

class LeaveViewSet(viewsets.ModelViewSet):
    queryset = Leave.objects.all()
    serializer_class = LeaveSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get_queryset(self):
        return Leave.objects.filter(company=self.request.user.company)

    def perform_create(self, serializer):
        company = self.request.user.company
        if not company:
            raise serializers.ValidationError("No company found for your admin.")
        serializer.save(company=company)

class LearningCornerViewSet(viewsets.ModelViewSet):
    queryset = LearningCorner.objects.all()
    serializer_class = LearningCornerSerializer
    permission_classes = [IsAuthenticated,IsAdminUser]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['title', 'description']

    def get_queryset(self):
        user = self.request.user
        return LearningCorner.objects.filter(company=user.company).order_by('id')

    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company)


class NotificationViewSet(viewsets.ModelViewSet):
    queryset = Notification.objects.all()
    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated,IsAdminUser]

    def get_queryset(self):
        user = self.request.user
        return Notification.objects.filter(company=user.company)

    def perform_create(self, serializer):
        # Ensure notifications are strictly tied to the admin's company.
        company = self.request.user.company
        if company:
            serializer.save(company=company)
        else:
            serializer.save()

class ShiftPolicyViewSet(viewsets.ModelViewSet):
    serializer_class = ShiftPolicySerializer
    permission_classes = [IsAuthenticated,IsAdminUser]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['shift_type']

    def get_queryset(self):
        return ShiftPolicy.objects.filter(company=self.request.user.company)

    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company)


class DepartmentWiseWorkingDaysViewSet(viewsets.ModelViewSet):
    queryset = DepartmentWiseWorkingDays.objects.all()
    serializer_class = DepartmentWiseWorkingDaysSerializer
    permission_classes = [IsAuthenticated,IsAdminUser]

    def get_queryset(self):
        user = self.request.user
        company = getattr(user, 'company', None)
        qs = DepartmentWiseWorkingDays.objects.all()
        if company:
            qs = qs.filter(company=company)
        return qs

    def _check_duplicate(self, department_id, shift_ids, company_id, current_id=None):
        existing_qs = DepartmentWiseWorkingDays.objects.filter(department_id=department_id)
        if company_id:
            existing_qs = existing_qs.filter(company_id=company_id)
        else:
            existing_qs = existing_qs.filter(company__isnull=True)
            
        if current_id:
            existing_qs = existing_qs.exclude(id=current_id)

        if not shift_ids:
            if existing_qs.filter(shifts__isnull=False).exists():
                return 'Shift-specific records already exist. Cannot add "All Shifts" record.'
        else:
            if existing_qs.filter(shifts__isnull=True).exists():
                return 'An "All Shifts" record exists. Cannot add shift-specific records.'

            # If any requested shift is already assigned in another record for this department
            conflicting = existing_qs.filter(shifts__id__in=shift_ids).distinct()
            if conflicting.exists():
                return 'One or more of the selected shifts are already configured for this department.'
        return None

    def create(self, request, *args, **kwargs):
        department_id = request.data.get('department')
        shift_ids = request.data.get('shifts', [])
        company_id = self.request.user.company.id

        if not department_id:
            return Response({'detail': 'Department is required.'}, status=400)

        err = self._check_duplicate(department_id, shift_ids, company_id)
        if err:
            return Response({'detail': err}, status=400)

        serializer = self.get_serializer(data={
            'department': department_id,
            'shifts': shift_ids,
            'working_days_count': request.data.get('working_days_count'),
            'week_start_day': request.data.get('week_start_day'),
            'week_end_day': request.data.get('week_end_day'),
            'working_days': request.data.get('working_days', []),
            'weekend_days': request.data.get('weekend_days', []),
            'company': company_id
        })
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        
        department_id = request.data.get('department', instance.department_id)
        shift_ids = request.data.get('shifts')
        if shift_ids is None:
            shift_ids = list(instance.shifts.values_list('id', flat=True))
            
        company_id = getattr(self.request.user.company, 'id', None)
        
        err = self._check_duplicate(department_id, shift_ids, company_id, current_id=instance.id)
        if err:
            return Response({'detail': err}, status=400)
            
        return super().update(request, *args, **kwargs)


class CalendarEventViewSet(viewsets.ModelViewSet):
    serializer_class = CalendarEventSerializer

    def get_permissions(self):
        if self.action in ("list", "retrieve"):
            return [IsAuthenticated(), CanReadCompanyCalendar()]
        return [IsAuthenticated(), IsAdminUser()]

    def get_queryset(self):
        return CalendarEvent.objects.filter(company=self.request.user.company)

    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company)

    @action(detail=False, methods=['post'], url_path='bulk-import')
    def bulk_import(self, request):
        """
        Bulk import company holidays/events.

        Expected JSON body:
        {
          "rows": [
            { "date": "2026-03-03", "name": "Holiday name", "description": "", "is_holiday": true }
          ]
        }
        """
        company = request.user.company
        rows = request.data.get('rows')
        if not isinstance(rows, list):
            return Response({'detail': 'rows must be a list'}, status=status.HTTP_400_BAD_REQUEST)

        created = 0
        updated = 0
        skipped = 0
        out = []

        for r in rows:
            if not isinstance(r, dict):
                skipped += 1
                continue
            name = (r.get('name') or '').strip()
            date_val = r.get('date')
            if not name or not date_val:
                skipped += 1
                continue
            desc = (r.get('description') or '').strip()
            is_holiday = bool(r.get('is_holiday', True))

            obj, was_created = CalendarEvent.objects.get_or_create(
                company=company,
                date=date_val,
                name=name,
                defaults={'description': desc, 'is_holiday': is_holiday},
            )
            if was_created:
                created += 1
            else:
                changed = False
                if desc and obj.description != desc:
                    obj.description = desc
                    changed = True
                if obj.is_holiday != is_holiday:
                    obj.is_holiday = is_holiday
                    changed = True
                if changed:
                    obj.save(update_fields=['description', 'is_holiday', 'updated_at'])
                    updated += 1
                else:
                    skipped += 1
            out.append({'id': obj.id, 'name': obj.name, 'date': str(obj.date), 'is_holiday': obj.is_holiday})

        return Response(
            {'created': created, 'updated': updated, 'skipped': skipped, 'results': out},
            status=status.HTTP_200_OK,
        )


class ChatConversationViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsCompanyChatUser]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ["name", "members__user__username", "members__user__email"]
    serializer_class = ChatConversationSerializer

    def get_queryset(self):
        user = self.request.user
        return (
            ChatConversation.objects.filter(company=user.company, members__user=user)
            .distinct()
            .order_by("-updated_at")
        )

    def perform_create(self, serializer):
        # Only group conversations created via REST; DMs can be created via action below
        user = self.request.user
        company = user.company
        conv_type = self.request.data.get("type") or "group"
        if conv_type != "group":
            raise serializers.ValidationError({"type": "Only group conversations can be created here."})
        name = (self.request.data.get("name") or "").strip()
        if not name:
            raise serializers.ValidationError({"name": "Group name is required."})

        conversation = serializer.save(company=company, created_by=user, type="group", name=name)
        # creator becomes admin with full permissions
        ChatConversationMember.objects.create(
            conversation=conversation,
            user=user,
            role="admin",
            can_add_members=True,
            can_remove_members=True,
            can_revoke_roles=True,
        )

        # optional initial members
        member_ids = self.request.data.get("member_ids") or []
        if isinstance(member_ids, str):
            member_ids = [x for x in member_ids.split(",") if x.strip()]
        for uid in member_ids:
            try:
                uid_int = int(uid)
            except Exception:
                continue
            if uid_int == user.id:
                continue
            target = UserRegister.objects.filter(id=uid_int, company=company).first()
            if not target:
                continue
            ChatConversationMember.objects.get_or_create(conversation=conversation, user=target, defaults={"role": "member"})

    @action(detail=False, methods=["post"], url_path="dm")
    def create_dm(self, request):
        """Create (or get) a DM conversation between current user and another user in same company."""
        user = request.user
        company = user.company
        other_id = request.data.get("user_id")
        if not other_id:
            return Response({"detail": "user_id is required."}, status=400)
        company_id = company.id
        other = UserRegister.objects.filter(
            Q(id=other_id),
            Q(company_id=company_id) | Q(employee__company_id=company_id),
            Q(role__in=["admin", "employee"]),
        ).first()
        if not other:
            return Response({"detail": "User not found in your company."}, status=404)
        if other.id == user.id:
            return Response({"detail": "Cannot DM yourself."}, status=400)

        # find existing dm with exactly these two members
        existing = (
            ChatConversation.objects.filter(company=company, type="dm", members__user=user)
            .filter(members__user=other)
            .distinct()
            .first()
        )
        if existing:
            ser = self.get_serializer(existing)
            return Response(ser.data, status=200)

        conv = ChatConversation.objects.create(company=company, type="dm", created_by=user, name=None)
        ChatConversationMember.objects.create(conversation=conv, user=user, role="member")
        ChatConversationMember.objects.create(conversation=conv, user=other, role="member")
        ser = self.get_serializer(conv)
        return Response(ser.data, status=201)

    @action(detail=True, methods=["post"], url_path="members/add")
    def add_member(self, request, pk=None):
        user = request.user
        conv = self.get_object()
        if conv.type != "group":
            return Response({"detail": "Members can only be managed for groups."}, status=400)
        actor = ChatConversationMember.objects.filter(conversation=conv, user=user).first()
        if not actor or not actor.can_add_members:
            return Response({"detail": "Not allowed."}, status=403)

        # Support adding a single user_id OR multiple member_ids
        ids = request.data.get("member_ids") or request.data.get("user_ids") or request.data.get("user_id")
        if isinstance(ids, str):
            ids = [x for x in ids.split(",") if x.strip()]
        if isinstance(ids, (int, float)):
            ids = [int(ids)]
        if not isinstance(ids, list):
            ids = [ids] if ids else []

        added = 0
        for member_id in ids:
            try:
                mid = int(member_id)
            except Exception:
                continue
            if mid == user.id:
                continue
            target = UserRegister.objects.filter(id=mid, company=user.company).first()
            if not target:
                continue
            ChatConversationMember.objects.get_or_create(conversation=conv, user=target, defaults={"role": "member"})
            added += 1
        conv.save(update_fields=["updated_at"])
        return Response({"detail": f"Members added: {added}."}, status=200)

    @action(detail=True, methods=["post"], url_path="members/remove")
    def remove_member(self, request, pk=None):
        user = request.user
        conv = self.get_object()
        if conv.type != "group":
            return Response({"detail": "Members can only be managed for groups."}, status=400)
        actor = ChatConversationMember.objects.filter(conversation=conv, user=user).first()
        if not actor or not actor.can_remove_members:
            return Response({"detail": "Not allowed."}, status=403)

        member_id = request.data.get("user_id")
        target_member = ChatConversationMember.objects.filter(conversation=conv, user_id=member_id).first()
        if not target_member:
            return Response({"detail": "Member not found."}, status=404)
        target_member.delete()
        conv.save(update_fields=["updated_at"])
        return Response({"detail": "Member removed."}, status=200)

    @action(detail=True, methods=["patch"], url_path="members/permissions")
    def update_member_permissions(self, request, pk=None):
        user = request.user
        conv = self.get_object()
        actor = ChatConversationMember.objects.filter(conversation=conv, user=user).first()
        if not actor or not actor.can_revoke_roles:
            return Response({"detail": "Not allowed."}, status=403)

        member_id = request.data.get("user_id")
        target = ChatConversationMember.objects.filter(conversation=conv, user_id=member_id).first()
        if not target:
            return Response({"detail": "Member not found."}, status=404)

        # Normalize roles and permissions:
        # - admin: full access
        # - member: message only (no member management)
        # - viewer: read-only (no sending messages enforced elsewhere)
        new_role = request.data.get("role", target.role)
        if new_role not in ("admin", "member", "viewer"):
            new_role = target.role

        if new_role == "admin":
            target.role = "admin"
            target.can_add_members = True
            target.can_remove_members = True
            target.can_revoke_roles = True
        elif new_role == "viewer":
            target.role = "viewer"
            target.can_add_members = False
            target.can_remove_members = False
            target.can_revoke_roles = False
        else:
            target.role = "member"
            target.can_add_members = False
            target.can_remove_members = False
            target.can_revoke_roles = False
        target.save()
        return Response({"detail": "Permissions updated."}, status=200)

    @action(detail=True, methods=["post"], url_path="seen")
    def mark_as_seen(self, request, pk=None):
        """Manually mark conversation as seen by the current user."""
        conv = self.get_object()
        ChatConversationMember.objects.filter(conversation=conv, user=request.user).update(last_seen_at=timezone.now())
        return Response({"detail": "Conversation marked as seen."}, status=200)


class ChatMessageViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsCompanyChatUser]
    pagination_class = CustomPagination
    serializer_class = ChatMessageSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        user = self.request.user
        qs = ChatMessage.objects.filter(company=user.company, conversation__members__user=user).distinct()
        conversation_id = self.request.query_params.get("conversation")
        if conversation_id:
            qs = qs.filter(conversation_id=conversation_id)
        return qs.order_by("-created_at")

    def list(self, request, *args, **kwargs):
        resp = super().list(request, *args, **kwargs)
        # Mark conversation as seen when messages are fetched.
        conversation_id = request.query_params.get("conversation")
        if conversation_id:
            ChatConversationMember.objects.filter(conversation_id=conversation_id, user_id=request.user.id).update(last_seen_at=timezone.now())
        return resp

    def perform_create(self, serializer):
        user = self.request.user
        # For multipart/form-data uploads, some clients omit/rename keys; prefer validated_data when present.
        conv_obj = serializer.validated_data.get("conversation") if hasattr(serializer, "validated_data") else None
        conv_id = (
            getattr(conv_obj, "id", None)
            or self.request.data.get("conversation")
            or self.request.data.get("conversation_id")
        )
        if not conv_id:
            raise serializers.ValidationError({"conversation": "conversation (or conversation_id) is required."})

        conv = ChatConversation.objects.filter(id=conv_id, company=user.company, members__user=user).first()
        if not conv:
            raise serializers.ValidationError({"conversation": "Conversation not found or not allowed."})
        if conv.type == "group":
            member = ChatConversationMember.objects.filter(conversation_id=conv.id, user_id=user.id).only("role").first()
            if member and member.role == "viewer":
                raise serializers.ValidationError({"detail": "View-only members cannot send messages."})
        msg = serializer.save(company=user.company, sender=user, conversation=conv)
        # Sender has "seen" their own message
        ChatConversationMember.objects.filter(conversation_id=conv.id, user_id=user.id).update(last_seen_at=timezone.now())
        # bump conversation updated_at
        ChatConversation.objects.filter(id=conv.id).update(updated_at=timezone.now())
        return msg


class ChatCompanyUsersAPIView(APIView):
    """
    List users in the current admin's company (admins + employees) for starting DMs.
    """

    permission_classes = [IsAuthenticated, IsCompanyChatUser]

    def _ensure_employee_profiles_for_company(self, company):
        """
        Ensure `UserRegister(role='employee')` users are represented in chat user picker
        even if they were created via SSO and don't yet have an Employee profile.
        """
        if not company:
            return

        missing_profile_users = (
            UserRegister.objects.filter(role__in=["employee", "admin"], company=company, employee__isnull=True)
            .only("id", "first_name", "last_name", "email", "company_id")
        )
        for u in missing_profile_users:
            if not u.email:
                continue
            existing = Employee.objects.filter(company=company, email__iexact=u.email).first()
            if existing:
                if not existing.user_id:
                    existing.user_id = u.id
                    if not existing.first_name:
                        existing.first_name = u.first_name or existing.first_name
                    if not existing.last_name:
                        existing.last_name = u.last_name or existing.last_name
                    if not existing.email:
                        existing.email = u.email
                    existing.save(update_fields=["user", "first_name", "last_name", "email"])
                continue
            Employee.objects.create(
                user_id=u.id,
                company=company,
                first_name=u.first_name or None,
                last_name=u.last_name or None,
                email=u.email,
            )

    def get(self, request):
        user = request.user
        company = getattr(user, "company", None)
        if not company:
            # fallback for employee accounts missing UserRegister.company
            emp = Employee.objects.filter(user=user).select_related("company").first()
            company = emp.company if emp else None
        if not company:
            return Response({"results": []}, status=200)

        # Keep chat picker complete for this company
        self._ensure_employee_profiles_for_company(company)

        q = (request.query_params.get("q") or "").strip()
        # Include:
        # - users where UserRegister.company == company
        # - employee-linked users where Employee.company == company
        company_id = company.id
        qs = (
            UserRegister.objects.filter(Q(company_id=company_id) | Q(employee__company_id=company_id))
            .exclude(id=user.id)
            .filter(role__in=["admin", "employee"])
            .select_related("employee")
            .distinct()
        )
        if q:
            qs = qs.filter(
                Q(username__icontains=q)
                | Q(email__icontains=q)
                | Q(first_name__icontains=q)
                | Q(last_name__icontains=q)
            )
        qs = qs.order_by("username")[:50]

        results = [
            {
                "id": u.id,
                "username": u.username,
                # Prefer Employee profile data when available (SSO/legacy users sometimes
                # have incorrect/missing fields on UserRegister).
                "email": (getattr(getattr(u, "employee", None), "email", None) or u.email or u.username),
                "first_name": (getattr(getattr(u, "employee", None), "first_name", None) or u.first_name),
                "last_name": (getattr(getattr(u, "employee", None), "last_name", None) or u.last_name),
                "role": u.role,
            }
            for u in qs
        ]
        return Response({"results": results}, status=200)

class RelievedEmployeeViewSet(viewsets.ModelViewSet):

    @action(detail=False, methods=['get'], url_path='search-employee')
    def search_employee(self, request):
        query = request.GET.get('q', '')
        company = request.user.company
        employees = Employee.objects.filter(
            is_active=True,
            company=company
        ).filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(employee_id__icontains=query)
        )[:20]
        results = [
            {
                'id': emp.id,
                'employee_id': emp.employee_id,
                'full_name': emp.full_name,
                'department': emp.department.department_name if emp.department else '',
                'designation': emp.designation.designation_name if emp.designation else '',
            }
            for emp in employees
        ]
        return Response(results, status=status.HTTP_200_OK)
    queryset = RelievedEmployee.objects.all()
    serializer_class = RelievedEmployeeSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = [
        'employee__employee_id',
        'employee__first_name',
        'employee__middle_name',
        'employee__last_name',
        'employee__full_name',
        'employee__department__department_name',
        'employee__designation__designation_name',
    ]

    def get_queryset(self):
        return RelievedEmployee.objects.filter(employee__company=self.request.user.company)

    def perform_create(self, serializer):
        
        relieved_instance = serializer.save()
       
        employee = relieved_instance.employee
        if employee:
           
            assigned_supply = EmployeeSupplyAssignment.objects.filter(employee=employee)
            for row in assigned_supply:
                si = row.supply_item
                si.available_quantity += 1
                si.save(update_fields=['available_quantity', 'updated_at'])
            assigned_supply.delete()
            # Mark employee as inactive
            employee.is_active = False
            employee.save()



class SalaryStructureViewSet(viewsets.ModelViewSet):
    serializer_class = SalaryStructureSerializer

    def get_queryset(self):
        return SalaryStructure.objects.filter(company=self.request.user.company).order_by('-created_at')
    
    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company)

class GrossSalaryComponentViewSet(viewsets.ModelViewSet):
    serializer_class = GrossSalaryComponentSerializer
    permission_classes = [IsAuthenticated, IsAdminUser | IsMaster]

    def get_queryset(self):
        return GrossSalaryComponent.objects.filter(company=self.request.user.company)

    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company)

class SalaryDeductionComponentViewSet(viewsets.ModelViewSet):
    serializer_class = SalaryDeductionComponentSerializer
    permission_classes = [IsAuthenticated, IsAdminUser | IsMaster]

    def get_queryset(self):
        return SalaryDeductionComponent.objects.filter(company=self.request.user.company)

    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company)

class PayrollBatchViewSet(viewsets.ModelViewSet):
    serializer_class = PayrollBatchSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    pagination_class = CustomPagination

    def get_queryset(self):
        return PayrollBatch.objects.filter(company=self.request.user.company).order_by('-year', '-month')

    @action(detail=True, methods=['post'], url_path='finalize')
    def finalize(self, request, pk=None):
        try:
            batch = self.get_object()

            if batch.company != request.user.company:
                return Response({'detail': 'Not allowed'}, status=status.HTTP_403_FORBIDDEN)

            if batch.status == 'Locked':
                return Response({'error': 'Batch already finalized.'}, status=400)

            salary_structure = SalaryStructure.objects.filter(company=batch.company).order_by('-created_at').first()
            if not salary_structure:
                return Response({'error': 'No Salary Structure found.'}, status=400)

            employees = Employee.objects.filter(company=batch.company)
            total_days = salary_structure.total_working_days or 30

            first_day = date(batch.year, batch.month, 1)
            last_day = date(batch.year, batch.month, calendar.monthrange(batch.year, batch.month)[1])

            payroll_data = []

            for emp in employees:
                gross = emp.gross_salary or Decimal(0)

                basic = gross * (salary_structure.basic_percent or 0) / 100
                hra = gross * (salary_structure.hra_percent or 0) / 100
                conveyance = gross * (salary_structure.conveyance_percent or 0) / 100
                medical = gross * (salary_structure.medical_percent or 0) / 100
                special = gross * (salary_structure.special_percent or 0) / 100
                service = gross * (salary_structure.service_charge_percent or 0) / 100

                per_day_salary = gross / total_days if total_days else Decimal(0)

                present_days = Attendance.objects.filter(
                    employee=emp,
                    date__range=(first_day, last_day)
                ).values('date').distinct().count()

                paid_leaves = EmpLeave.objects.filter(
                    employee=emp,
                    leave_type__is_paid=True,
                    status='Approved',
                    from_date__gte=first_day,
                    to_date__lte=last_day
                ).count()

                lop_days = EmpLeave.objects.filter(
                    employee=emp,
                    leave_type__is_paid=False,
                    status='Approved',
                    from_date__gte=first_day,
                    to_date__lte=last_day
                ).count()

                days_paid = present_days + paid_leaves
                adjusted_gross = per_day_salary * Decimal(days_paid)

                pf = basic * Decimal('0.12')

                tax_slab = IncomeTaxConfig.objects.filter(
                    company=batch.company,
                    salary_from__lte=gross,
                    salary_to__gte=gross
                ).first()

                income_tax = gross * (tax_slab.tax_percent / Decimal('100')) if tax_slab else Decimal(0)

                extra_allowances = sum(a.amount for a in salary_structure.allowances.all()) or Decimal(0)
                extra_deductions = sum(d.amount for d in salary_structure.deductions.all()) or Decimal(0)

                net_pay = adjusted_gross + extra_allowances - (pf + income_tax + extra_deductions)
                net_pay = max(net_pay, Decimal(0))

                payroll = Payroll.objects.create(
                    batch=batch,
                    company=batch.company,
                    employee=emp,
                    salary_structure=salary_structure,
                    gross_salary=gross,
                    basic_salary=basic,
                    hra=hra,
                    conveyance=conveyance,
                    medical=medical,
                    special_allowance=special,
                    service_charges=service,
                    pf=pf,
                    net_pay=net_pay,
                    total_working_days=total_days,
                    days_paid=days_paid,
                    loss_of_pay_days=lop_days,
                    income_tax=income_tax,
                    payroll_date=timezone.now().date(),
                )

                payroll_data.append(PayrollSerializer(payroll).data)

            batch.status = 'Locked'
            batch.save()

            return Response({
                'message': f'Payroll batch {batch.id} finalized.',
                'batch': PayrollBatchSerializer(batch).data,
                'payrolls': payroll_data
            })

        except PayrollBatch.DoesNotExist:
            return Response({'error': 'Batch not found'}, status=status.HTTP_404_NOT_FOUND)
        
        return Response({'message': 'Payslips sent to all employees.'})

    @action(detail=True, methods=['post'], url_path='rollout-payslips')
    def rollout_payslips(self, request, pk=None):
        batch = self.get_object()
        if batch.status != 'Locked':
            return Response({'error': 'Batch must be locked before rolling out payslips.'}, status=400)

        company = batch.company
        logo_path = company.logo.path if company.logo and hasattr(company.logo, 'path') else None

        # Filter employees if provided (optional)
        employee_ids = request.data.get('employee_ids', None)
        payrolls = Payroll.objects.filter(batch=batch)
        if employee_ids:
            payrolls = payrolls.filter(employee_id__in=employee_ids)

        count = 0
        for payroll in payrolls:
            employee = payroll.employee
            
            # Generate Unique Payslip ID
            payslip_id = f"PS-{batch.year}{str(batch.month).zfill(2)}-{employee.employee_id}-{uuid.uuid4().hex[:6].upper()}"
            
            # Use transaction or update_or_create to avoid duplicates
            payslip, created = Payslip.objects.get_or_create(
                payroll=payroll,
                defaults={
                    'employee': employee,
                    'company': company,
                    'payslip_id': payslip_id,
                    'month': batch.month,
                    'year': batch.year,
                    'status': 'Draft'
                }
            )

            if created or not payslip.file:
                # Generate PDF with QR Code and ID
                pdf_buffer = generate_payslip_pdf(employee, payroll, batch, company=company, logo_path=logo_path, payslip_id=payslip.payslip_id)
                filename = f"Payslip_{payslip.payslip_id}.pdf"
                payslip.file.save(filename, ContentFile(pdf_buffer.read()), save=True)
                count += 1

        return Response({'message': f'Payslips rolled out to {count} employees.'})

           
class GeneratePayrollView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def post(self, request):
        company = request.user.company
        today = timezone.now().date()
        month = today.month
        year = today.year

        # Prevent duplicate finalized batch
        if PayrollBatch.objects.filter(company=company, month=month, year=year, status='Locked').exists():
            return Response({'error': 'Finalized batch already exists for this month'}, status=400)

        # Create or get draft batch
        batch, _ = PayrollBatch.objects.get_or_create(
            company=company,
            month=month,
            year=year,
            defaults={'status': 'Draft'}
        )

        salary_structure = SalaryStructure.objects.filter(company=company).order_by('-created_at').first()
        if not salary_structure:
            return Response({'error': 'No Salary Structure found.'}, status=400)

        employees = Employee.objects.filter(company=company)
        total_days = salary_structure.total_working_days or 30

        first_day = today.replace(day=1)
        last_day = today.replace(day=calendar.monthrange(year, month)[1])

        preview_data = []

        for emp in employees:
            gross = emp.gross_salary or Decimal(0)

            basic = gross * (salary_structure.basic_percent or 0) / 100
            hra = gross * (salary_structure.hra_percent or 0) / 100
            conveyance = gross * (salary_structure.conveyance_percent or 0) / 100
            medical = gross * (salary_structure.medical_percent or 0) / 100
            special = gross * (salary_structure.special_percent or 0) / 100
            service = gross * (salary_structure.service_charge_percent or 0) / 100

            per_day_salary = gross / total_days if total_days else Decimal(0)

            present_days = Attendance.objects.filter(
                employee=emp,
                date__range=(first_day, last_day)
            ).values('date').distinct().count()

            paid_leaves = EmpLeave.objects.filter(
                employee=emp,
                leave_type__is_paid=True,
                status='Approved',
                from_date__gte=first_day,
                to_date__lte=last_day
            ).count()

            lop_days = EmpLeave.objects.filter(
                employee=emp,
                leave_type__is_paid=False,
                status='Approved',
                from_date__gte=first_day,
                to_date__lte=last_day
            ).count()

            days_paid = present_days + paid_leaves
            adjusted_gross = per_day_salary * Decimal(days_paid)

            pf = basic * Decimal('0.12')

            tax_slab = IncomeTaxConfig.objects.filter(
                company=company,
                salary_from__lte=gross,
                salary_to__gte=gross
            ).first()

            income_tax = gross * (tax_slab.tax_percent / Decimal('100')) if tax_slab else Decimal(0)
            
            # Loan EMI deduction (Period-aware: only deduct if repayment day falls in this period)
            loan_emi = Decimal(0)
            loans = LoanApplication.objects.filter(employee=emp.user, status='APPROVED')
            
            from datetime import timedelta
            import calendar
            
            for l in loans:
                start_dt = l.created_at.date()
                # Calculate end date of tenure
                m_total = start_dt.month + l.repayment_months
                e_year = start_dt.year + (m_total - 1) // 12
                e_month = (m_total - 1) % 12 + 1
                _, e_last = calendar.monthrange(e_year, e_month)
                end_dt = date(e_year, e_month, e_last)
                
                # Skip if loan period has ended
                if first_day > end_dt:
                    continue
                
                r_day = start_dt.day
                curr = first_day
                match = False
                while curr <= last_day:
                    _, lday = calendar.monthrange(curr.year, curr.month)
                    is_m = (curr.day == r_day)
                    if not is_m and r_day > lday and curr.day == lday:
                        is_m = True
                    if is_m and curr <= end_dt:
                        loan_emi += Decimal(str(l.emi_amount))
                    curr += timedelta(days=1)
            # Loan Disbursement (Principal if approved in this month)
            loan_disb = LoanApplication.objects.filter(
                employee=emp.user,
                status__in=['APPROVED', 'CLEARED'],
                created_at__date__range=(first_day, last_day)
            ).aggregate(total=Sum('requested_amount'))['total'] or Decimal(0)
            loan_disb = Decimal(str(loan_disb))

            extra_allowances = sum(a.amount for a in salary_structure.allowances.all()) or Decimal(0)
            extra_deductions = sum(d.amount for d in salary_structure.deductions.all()) or Decimal(0)

            net_pay = adjusted_gross + extra_allowances + loan_disb - (pf + income_tax + extra_deductions + loan_emi)
            net_pay = max(net_pay, Decimal(0))

            # Create unsaved Payroll instance
            fake_payroll = Payroll(
                batch=batch,
                company=company,
                employee=emp,
                salary_structure=salary_structure,
                gross_salary=gross,
                basic_salary=basic,
                hra=hra,
                conveyance=conveyance,
                medical=medical,
                special_allowance=special,
                service_charges=service,
                pf=pf,
                income_tax=income_tax,
                loan_emi=loan_emi,
                loan_disbursement=loan_disb,
                net_pay=net_pay,
                total_working_days=total_days,
                days_paid=days_paid,
                loss_of_pay_days=lop_days,
                payroll_date=timezone.now().date(),
            )

            # Serialize without saving
            serializer = PayrollSerializer(fake_payroll)
            preview_data.append(serializer.data)

        return Response({
            'batch_id': batch.id,
            'batch_status': batch.status,
            'payroll_preview': preview_data
        })



class PayrollViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = PayrollSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get_queryset(self):
        user_company = self.request.user.company
        queryset = Payroll.objects.filter(company=user_company)

        batch_id = self.request.query_params.get('batch_id')
        year = self.request.query_params.get('year')
        month = self.request.query_params.get('month')
        day = self.request.query_params.get('day')

        if batch_id:
            queryset = queryset.filter(batch_id=batch_id)
        if year:
            queryset = queryset.filter(payroll_date__year=year)
        if month:
            queryset = queryset.filter(payroll_date__month=month)
        if day:
            queryset = queryset.filter(payroll_date__day=day)

        return queryset.order_by('-payroll_date')

class PayslipViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = PayslipSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['payslip_id', 'employee__first_name', 'employee__last_name', 'employee__employee_id']

    def get_queryset(self):
        user = self.request.user
        if user.role in ['admin', 'master']:
            return Payslip.objects.filter(company=user.company).order_by('-year', '-month')
        return Payslip.objects.filter(employee__user=user, status='Published').order_by('-year', '-month')

    @action(detail=False, methods=['post'], url_path='generate')
    def generate(self, request):
        payroll_id = request.data.get('payroll_id')
        is_report = request.data.get('is_report', False)
        regenerate = request.data.get('regenerate', False)
        start_date_req = request.data.get('start_date')
        end_date_req = request.data.get('end_date')
        
        try:
            if is_report:
                record = FinalizedSalary.objects.get(id=payroll_id)
                employee = record.employee
                
                # If dates are provided, find the correct record for those dates
                # instead of blindly updating (which can violate unique_together)
                if start_date_req and end_date_req:
                    req_from = datetime.strptime(start_date_req, '%Y-%m-%d').date()
                    req_to = datetime.strptime(end_date_req, '%Y-%m-%d').date()
                    
                    # Check if a different record already exists for these dates
                    existing = FinalizedSalary.objects.filter(
                        employee=employee, from_date=req_from, to_date=req_to
                    ).exclude(id=record.id).first()
                    
                    if existing:
                        # Use the existing record for those dates instead
                        record = existing
                    elif record.from_date != req_from or record.to_date != req_to:
                        # Only update dates if they actually changed and no conflict
                        record.from_date = req_from
                        record.to_date = req_to
                        record.save()
                
                month = record.from_date.month
                year = record.from_date.year
            else:
                record = Payroll.objects.get(id=payroll_id)
                employee = record.employee
                month = record.batch.month
                year = record.batch.year
                
            # Find existing payslip: try reverse FK first, then fallback to employee+month+year lookup
            payslip = None
            try:
                payslip = record.payslip_record
            except Exception:
                pass
            
            if not payslip:
                if not is_report:
                    # Fallback: find by employee + month + year for full month payrolls ONLY
                    # We don't do this for is_report because multiple date ranges can exist in the same month
                    payslip = Payslip.objects.filter(
                        employee=employee, month=month, year=year, company=request.user.company
                    ).first()
                    
                    # If found, link it to the current record for future lookups
                    if payslip and not payslip.payroll_id:
                        payslip.payroll = record
                        payslip.save(update_fields=['payroll'])
            
            if payslip and not regenerate and not is_report:
                return Response({'detail': 'Payslip already generated', 'file': payslip.file.url}, status=200)

            # Always recalculate from live attendance for FinalizedSalary reports
            # This ensures the payslip matches the PayrollReport calculations
            if is_report:
                m = compute_attendance_metrics(employee, record.from_date, record.to_date)
                # Recompute days_paid
                d_paid = Decimal(str(m['present_days'])) + (Decimal(str(m['half_days'])) * Decimal('0.5')) + Decimal(str(m['paid_leaves']))
                record.days_paid = d_paid
                
                # Update config with fresh metrics
                conf = record.config or {}
                conf.update({
                    'present_days': m['present_days'],
                    'half_days': m['half_days'],
                    'paid_leaves': m['paid_leaves'],
                    'unpaid_leaves': m['unpaid_leaves'],
                    'expected_working_days': m['expected_working_days'],
                    'absent_days': m['absent_days'],
                    'overtime_hours': m['overtime_hours'],
                    'checked_in_days': m['checked_in_days'],
                })
                record.config = conf
                
                # Recompute earned basic based on new days_paid
                fixed_basic = employee.basic_salary or Decimal(0)
                if m['expected_working_days'] > 0:
                    record.earned_basic = ((fixed_basic / Decimal(m['expected_working_days'])) * d_paid).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                
                # Refresh Loans and Assets
                record.loan_emi = compute_loan_emi(employee, record.from_date, record.to_date)
                record.loan_disbursement = compute_loan_disbursement(employee, record.from_date, record.to_date)
                record.asset_deduction = compute_asset_deductions(employee, record.from_date, record.to_date)
                
                # Recalculate OT pay from fresh attendance data
                ot_hours = Decimal(str(m['overtime_hours']))
                record.ot_hours = ot_hours
                if conf.get('otEnabled') and m['expected_working_days'] > 0:
                    hourly_rate = (fixed_basic / Decimal(str(m['expected_working_days']))) / Decimal(8)
                    record.ot_pay = (hourly_rate * ot_hours).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                else:
                    record.ot_pay = Decimal(0)
                
                # Recalculate total_gross and net_salary for consistency
                total_gross = record.earned_basic + record.ot_pay + record.loan_disbursement
                
                # Add dynamic components — respect saved gChk config from PayrollReport
                from app.models import GrossSalaryComponent, SalaryDeductionComponent
                all_gross = GrossSalaryComponent.objects.filter(company=employee.company, is_active=True)
                g_chk = conf.get('gChk', {})
                
                all_gross_names = set()
                for gc in all_gross:
                    all_gross_names.add(gc.name.lower())
                    # Respect saved config: skip if explicitly unchecked, default True for new components
                    if not g_chk.get(f'g-{gc.id}', True): continue
                    amount = (record.earned_basic * gc.value) / Decimal(100) if gc.calc_type == 'percentage' else gc.value
                    total_gross += amount
                
                # Add Salary Structure allowances (only items NOT covered by any GrossSalaryComponent)
                struct_allowances = compute_salary_structure_allowances(employee, record.earned_basic, total_gross)
                for sa in struct_allowances:
                    if sa['name'].lower() not in all_gross_names:
                        total_gross += sa['amount']
                
                # Calculate reimbursement from database
                from app.models import ReimbursementRequest
                reimbursements = ReimbursementRequest.objects.filter(
                    employee=employee,
                    status='approved',
                    created_at__date__range=[record.from_date, record.to_date]
                )
                total_reimb = sum(r.amount for r in reimbursements)
                
                record.total_gross = (total_gross + Decimal(str(total_reimb))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                
                # Total deductions
                total_ded = record.loan_emi or Decimal(0)
                total_ded += record.asset_deduction or Decimal(0)
                all_ded = SalaryDeductionComponent.objects.filter(company=employee.company, is_active=True)
                d_chk = conf.get('dChk', {})
                for dc in all_ded:
                    if not d_chk.get(f'd-{dc.id}', True): continue
                    
                    fixed_gross = employee.gross_salary or Decimal(0)
                    t_base = fixed_basic if dc.threshold_on == 'basic' else fixed_gross
                    if dc.has_threshold and t_base < dc.threshold_amount: continue
                    
                    d_base = record.earned_basic if dc.deduct_from == 'basic' else record.total_gross
                    amount = (d_base * dc.value) / Decimal(100) if dc.calc_type == 'percentage' else dc.value
                    total_ded += amount
                
                record.total_deductions = total_ded.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                
                record.net_salary = (record.total_gross - record.total_deductions).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                
                record.save()
            
            # Generate unique payslip ID or reuse existing
            unique_id = payslip.payslip_id if payslip else f"PAY-{uuid.uuid4().hex[:8].upper()}"
            
            # Generate PDF
            from .utils import generate_payslip_pdf
            pdf_file = generate_payslip_pdf(
                employee=employee,
                payroll=record if not is_report else None,
                batch=record.batch if not is_report else None,
                company=request.user.company,
                payslip_id=unique_id,
                finalized_salary=record if is_report else None
            )
            
            # Create or Update Payslip record
            if not payslip:
                payslip = Payslip.objects.create(
                    employee=employee,
                    company=request.user.company,
                    payroll=record if not is_report else None,
                    finalized_salary=record if is_report else None,
                    payslip_id=unique_id,
                    month=month,
                    year=year,
                    status='Draft'
                )
            
            # Delete old file first, then save new one (prevents browser caching old URL)
            if payslip.file:
                try:
                    payslip.file.delete(save=False)
                except Exception:
                    pass
            
            # Append timestamp to filename to bust browser cache
            ts = int(timezone.now().timestamp())
            payslip.file.save(f"payslip_{unique_id}_{ts}.pdf", ContentFile(pdf_file.read()))
            payslip.save()
            
            return Response({
                'message': 'Payslip generated successfully',
                'file': payslip.file.url,
                'payslip_id': payslip.payslip_id
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=400)

    @action(detail=False, methods=['get'], url_path='rollout-dashboard')
    def rollout_dashboard(self, request):
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        company = request.user.company
        
        # Parse dates
        try:
            if start_date and end_date:
                s_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
                e_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
            else:
                month = request.query_params.get('month')
                year = request.query_params.get('year')
                if month and year:
                    m, y = int(month), int(year)
                else:
                    today = timezone.now().date()
                    m, y = today.month, today.year
                s_dt = date(y, m, 1)
                e_dt = date(y, m, calendar.monthrange(y, m)[1])
        except Exception:
            today = timezone.now().date()
            s_dt = today.replace(day=1)
            e_dt = today

        employees = Employee.objects.filter(
            company=company, is_active=True, user__role='employee'
        ).select_related('department', 'shift_assigned')
        
        # Fetch all FinalizedSalary records for this period
        finalized_salaries = FinalizedSalary.objects.filter(
            company=company, from_date__gte=s_dt, to_date__lte=e_dt
        )
        final_map = {fs.employee_id: fs for fs in finalized_salaries}
        
        # Fetch payroll batch records
        payrolls = Payroll.objects.filter(company=company).filter(
            Q(batch__year__gt=s_dt.year) | Q(batch__year=s_dt.year, batch__month__gte=s_dt.month)
        ).filter(
            Q(batch__year__lt=e_dt.year) | Q(batch__year=e_dt.year, batch__month__lte=e_dt.month)
        )
        payroll_map = {p.employee_id: p for p in payrolls}
        
        # Fetch payslips linked to finalized salaries or payrolls
        payslip_by_fs = {}
        payslip_by_payroll = {}
        all_payslips = Payslip.objects.filter(company=company)
        for ps in all_payslips:
            if ps.finalized_salary_id:
                payslip_by_fs[ps.finalized_salary_id] = ps
            if ps.payroll_id:
                payslip_by_payroll[ps.payroll_id] = ps
        
        # Preload company-wide data
        holidays = set(CalendarEvent.objects.filter(
            company=company, is_holiday=True, date__gte=s_dt, date__lte=e_dt
        ).values_list('date', flat=True))
        
        # Preload all attendance records for the range
        all_attendance = Attendance.objects.filter(
            company=company, date__gte=s_dt, date__lte=e_dt
        ).select_related('employee', 'leave', 'leave__leave_type')
        att_by_emp = {}
        for att in all_attendance:
            att_by_emp.setdefault(att.employee_id, {})[att.date] = att
        
        # Preload all approved leaves for the range
        all_leaves = EmpLeave.objects.filter(
            company=company, status='Approved',
            from_date__lte=e_dt, to_date__gte=s_dt
        ).select_related('leave_type', 'employee')
        leaves_by_emp = {}
        for lv in all_leaves:
            leaves_by_emp.setdefault(lv.employee_id, []).append(lv)
        
        # Preload dept working days config
        dept_configs = {}
        for cfg in DepartmentWiseWorkingDays.objects.filter(company=company):
            dept_configs[cfg.department_id] = cfg
        
        # Preload dynamic salary components
        gross_comps = list(GrossSalaryComponent.objects.filter(company=company, is_active=True))
        deduction_comps = list(SalaryDeductionComponent.objects.filter(company=company, is_active=True))
        
        # Preload reimbursements
        from app.models import ReimbursementRequest
        reimb_by_emp = {}
        for r in ReimbursementRequest.objects.filter(
            company=company, status='approved',
            created_at__date__gte=s_dt, created_at__date__lte=e_dt
        ):
            reimb_by_emp.setdefault(r.employee_id, Decimal(0))
            reimb_by_emp[r.employee_id] += r.amount
        
        # Preload asset deductions
        from app.models import AssetRequest
        asset_ded_by_emp = {}
        for ar in AssetRequest.objects.filter(
            company=company, approval_status='approved',
            updated_at__date__range=[s_dt, e_dt]
        ).select_related('related_fixed_asset', 'related_supply_item'):
            emp_id = ar.requested_by_id
            if ar.related_supply_item:
                amt = (ar.related_supply_item.unit_price or Decimal(0)) * (ar.requested_quantity or 1)
            else:
                amt = ar.employee_payment_amount or Decimal(0)
            asset_ded_by_emp.setdefault(emp_id, Decimal(0))
            asset_ded_by_emp[emp_id] += amt
        
        # Preload loan data
        loan_emi_by_emp = {}
        loans = LoanApplication.objects.filter(
            employee__company=company, status__in=['APPROVED', 'CLEARED']
        )
        for l in loans:
            emp_profile = getattr(l.employee, 'employee_profile', None)
            if not emp_profile:
                continue
            emp_id = emp_profile.id
            start_dt = l.created_at.date()
            r_day = start_dt.day
            m_total = start_dt.month + l.repayment_months
            e_year = start_dt.year + (m_total - 1) // 12
            e_month = (m_total - 1) % 12 + 1
            _, e_last = calendar.monthrange(e_year, e_month)
            end_dt_loan = date(e_year, e_month, e_last)
            if s_dt > end_dt_loan:
                continue
            curr = s_dt
            while curr <= e_dt:
                _, lday = calendar.monthrange(curr.year, curr.month)
                is_m = (curr.day == r_day)
                if not is_m and r_day > lday and curr.day == lday:
                    is_m = True
                if is_m and curr <= end_dt_loan:
                    loan_emi_by_emp.setdefault(emp_id, Decimal(0))
                    loan_emi_by_emp[emp_id] += Decimal(str(l.emi_amount))
                curr += timedelta(days=1)

        data = []
        for emp in employees:
            fs = final_map.get(emp.id)
            p = payroll_map.get(emp.id)
            
            # ---- Compute attendance (same logic as PayrollAttendanceSummaryView) ----
            cfg = dept_configs.get(emp.department_id)
            weekend_days = [d.lower() for d in cfg.weekend_days] if cfg and cfg.weekend_days else ["saturday", "sunday"]
            
            shift = emp.shift_assigned
            s_full = shift.full_day_hours() if shift else 8.0
            s_half = shift.half_day_hours() if shift else 4.0
            
            emp_att_map = att_by_emp.get(emp.id, {})
            emp_leaves = leaves_by_emp.get(emp.id, [])
            
            # Pre-calculate leave dates
            leave_dates_paid = set()
            leave_dates_unpaid = set()
            for lv in emp_leaves:
                l_start = max(lv.from_date, s_dt)
                l_end = min(lv.to_date, e_dt)
                d = l_start
                while d <= l_end:
                    if lv.leave_type and lv.leave_type.is_paid:
                        leave_dates_paid.add(d)
                    else:
                        leave_dates_unpaid.add(d)
                    d += timedelta(days=1)
            
            present_days = 0
            half_days = 0
            absent_days = 0
            expected_working_days = 0
            total_overtime_seconds = 0
            checked_in_days = 0
            
            curr_date = s_dt
            while curr_date <= e_dt:
                day_name = curr_date.strftime('%A').lower()
                is_holiday = curr_date in holidays
                is_weekend = day_name in weekend_days
                is_work_day = not is_holiday and not is_weekend
                
                if is_work_day:
                    expected_working_days += 1
                
                att_rec = emp_att_map.get(curr_date)
                if att_rec:
                    if is_holiday or is_weekend:
                        if att_rec.total_work_duration:
                            total_overtime_seconds += att_rec.total_work_duration.total_seconds()
                        elif att_rec.check_in and att_rec.check_out:
                            total_overtime_seconds += (att_rec.check_out - att_rec.check_in).total_seconds()
                    else:
                        if att_rec.leave:
                            pass  # handled by leave_dates
                        elif att_rec.check_in and not att_rec.check_out:
                            checked_in_days += 1
                        else:
                            w_hours = 0.0
                            if att_rec.total_work_duration:
                                w_hours = att_rec.total_work_duration.total_seconds() / 3600
                            elif att_rec.check_in and att_rec.check_out:
                                w_hours = (att_rec.check_out - att_rec.check_in).total_seconds() / 3600
                            
                            if w_hours >= s_full:
                                present_days += 1
                            elif w_hours >= s_half:
                                half_days += 1
                            else:
                                absent_days += 1
                        
                        if att_rec.overtime_duration:
                            total_overtime_seconds += att_rec.overtime_duration.total_seconds()
                elif curr_date in leave_dates_paid or curr_date in leave_dates_unpaid:
                    pass
                elif is_work_day:
                    absent_days += 1
                
                curr_date += timedelta(days=1)
            
            paid_leaves = len(leave_dates_paid)
            unpaid_leaves = len(leave_dates_unpaid)
            overtime_hours = round(total_overtime_seconds / 3600, 2)
            
            # ---- Compute salary (same logic as PayrollReport frontend) ----
            basic = Decimal(str(emp.basic_salary or 0))
            gross_fixed = Decimal(str(emp.gross_salary or 0))
            
            if expected_working_days > 0:
                payable_days = Decimal(str(present_days)) + (Decimal(str(half_days)) * Decimal('0.5')) + Decimal(str(paid_leaves))
                earned_basic = (basic / Decimal(str(expected_working_days))) * payable_days
            else:
                payable_days = Decimal(0)
                earned_basic = basic
            
            earned_basic = earned_basic.quantize(Decimal('0.01'))
            
            # OT pay
            ot_pay = Decimal(0)
            if fs and fs.config and fs.config.get('otEnabled') and expected_working_days > 0:
                hourly_rate = (basic / Decimal(str(expected_working_days))) / Decimal(8)
                ot_pay = (hourly_rate * Decimal(str(overtime_hours))).quantize(Decimal('0.01'))
            
            # Gross components
            total_gross = earned_basic + ot_pay
            if fs and fs.config:
                g_chk = fs.config.get('gChk', {})
            else:
                g_chk = {f'g-{gc.id}': True for gc in gross_comps}
            
            for gc in gross_comps:
                if not g_chk.get(f'g-{gc.id}', True):
                    continue
                if gc.calc_type == 'percentage':
                    amt = (earned_basic * gc.value) / Decimal(100)
                else:
                    amt = gc.value
                total_gross += amt.quantize(Decimal('0.01'))
            
            # Add reimbursement
            reimb = reimb_by_emp.get(emp.id, Decimal(0))
            total_gross += reimb
            
            # Deduction components
            total_deductions = Decimal(0)
            if fs and fs.config:
                d_chk = fs.config.get('dChk', {})
            else:
                d_chk = {}
                for dc in deduction_comps:
                    if dc.has_threshold:
                        t_base = basic if dc.threshold_on == 'basic' else gross_fixed
                        d_chk[f'd-{dc.id}'] = float(t_base) >= float(dc.threshold_amount)
                    else:
                        d_chk[f'd-{dc.id}'] = True
            
            for dc in deduction_comps:
                if not d_chk.get(f'd-{dc.id}', True):
                    continue
                d_base = earned_basic if dc.deduct_from == 'basic' else total_gross
                if dc.calc_type == 'percentage':
                    amt = (d_base * dc.value) / Decimal(100)
                else:
                    amt = dc.value
                total_deductions += amt.quantize(Decimal('0.01'))
            
            # Add loan EMI
            emp_loan_emi = loan_emi_by_emp.get(emp.id, Decimal(0))
            total_deductions += emp_loan_emi
            
            # Add asset deduction
            emp_asset_ded = asset_ded_by_emp.get(emp.id, Decimal(0))
            total_deductions += emp_asset_ded
            
            net_pay = total_gross - total_deductions
            if net_pay < 0:
                net_pay = Decimal(0)
            
            # ---- Determine payslip record ----
            # Prefer FinalizedSalary, then Payroll
            # If neither exists, auto-create a FinalizedSalary so Generate works for any dates
            if not fs and not p:
                fs, _created = FinalizedSalary.objects.update_or_create(
                    employee=emp,
                    company=company,
                    from_date=s_dt,
                    to_date=e_dt,
                    defaults={
                        'basic_salary': basic,
                        'earned_basic': earned_basic,
                        'ot_pay': ot_pay,
                        'ot_hours': Decimal(str(overtime_hours)),
                        'total_gross': total_gross.quantize(Decimal('0.01')),
                        'total_deductions': total_deductions.quantize(Decimal('0.01')),
                        'loan_emi': emp_loan_emi,
                        'loan_disbursement': Decimal(0),
                        'asset_deduction': emp_asset_ded,
                        'net_salary': net_pay.quantize(Decimal('0.01')),
                        'days_paid': payable_days,
                        'config': {
                            'gChk': g_chk,
                            'dChk': d_chk,
                            'otEnabled': False,
                            'present_days': present_days,
                            'half_days': half_days,
                            'paid_leaves': paid_leaves,
                            'unpaid_leaves': unpaid_leaves,
                            'expected_working_days': expected_working_days,
                            'absent_days': absent_days,
                            'overtime_hours': overtime_hours,
                            'checked_in_days': checked_in_days,
                        }
                    }
                )
                final_map[emp.id] = fs
            
            main_record = fs if fs else p
            is_report = bool(fs)
            
            if fs:
                payslip = payslip_by_fs.get(fs.id)
            elif p:
                payslip = payslip_by_payroll.get(p.id)
            else:
                payslip = None
            
            # For payroll_id: use fs.id if report, else p.id
            payroll_id = fs.id if fs else (p.id if p else None)
            
            data.append({
                'employee_id_str': emp.employee_id,
                'employee_name': emp.full_name,
                'payroll_id': payroll_id,
                'batch_id': p.batch.id if p else None,
                'is_report': is_report,
                'net_pay': float(net_pay.quantize(Decimal('0.01'))),
                'details': {
                    'gross': float(total_gross.quantize(Decimal('0.01'))),
                    'deductions': float(total_deductions.quantize(Decimal('0.01'))),
                    'ot_pay': float(ot_pay),
                    'loan_emi': float(emp_loan_emi),
                    'asset_deduction': float(emp_asset_ded),
                    'days_paid': float(payable_days),
                    'present_days': present_days,
                    'half_days': half_days,
                    'paid_leaves': paid_leaves,
                    'unpaid_leaves': unpaid_leaves,
                    'expected_working_days': expected_working_days,
                    'absent_days': absent_days,
                    'overtime_hours': overtime_hours,
                    'checked_in_days': checked_in_days,
                    'earned_basic': float(earned_basic),
                    'basic_salary': float(basic),
                    'reimbursement': float(reimb),
                },
                'payslip_id': payslip.payslip_id if payslip else None,
                'payslip_status': payslip.status if payslip else 'Not Generated',
                'file': payslip.file.url if payslip and payslip.file else None,
                'id': payslip.id if payslip else None,
            })
        return Response(data)

    @action(detail=True, methods=['post'], url_path='publish')
    def publish(self, request, pk=None):
        payslip = self.get_object()
        payslip.status = 'Published'
        payslip.save()
        return Response({'message': 'Payslip published successfully'})

    @action(detail=False, methods=['post'], url_path='bulk-publish')
    def bulk_publish(self, request):
        ids = request.data.get('ids', [])
        Payslip.objects.filter(id__in=ids, company=request.user.company).update(status='Published')
        return Response({'message': f'{len(ids)} payslips published.'})

class IncomeTaxConfigViewSet(viewsets.ModelViewSet):
    serializer_class = IncomeTaxConfigSerializer

    def get_queryset(self):
        return IncomeTaxConfig.objects.filter(company=self.request.user.company)


class FinalizedSalaryViewSet(viewsets.ModelViewSet):
    queryset = FinalizedSalary.objects.all()
    serializer_class = FinalizedSalarySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = self.queryset.filter(company=self.request.user.company)
        employee = self.request.query_params.get('employee')
        from_date = self.request.query_params.get('from_date')
        to_date = self.request.query_params.get('to_date')
        if employee:
            qs = qs.filter(employee_id=employee)
        if from_date:
            qs = qs.filter(from_date=from_date)
        if to_date:
            qs = qs.filter(to_date=to_date)
        return qs

    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company)

class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    pagination_class = CustomPagination

    def get_queryset(self):
        # Base queryset for existing attendance records
        queryset = Attendance.objects.filter(company=self.request.user.company)
        
        from_date = self.request.query_params.get('from_date')
        to_date = self.request.query_params.get('to_date')
        if from_date:
            queryset = queryset.filter(date__gte=from_date)
        if to_date:
            queryset = queryset.filter(date__lte=to_date)
            
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(employee__employee_id__icontains=search) |
                Q(employee__first_name__icontains=search) |
                Q(employee__last_name__icontains=search)
            )
        return queryset

    def list(self, request, *args, **kwargs):
        company = request.user.company
        from_date_str = request.query_params.get('from_date')
        to_date_str = request.query_params.get('to_date')
        search = request.query_params.get('search', '').strip()
        status_filter = request.query_params.get('status', 'All').lower()
        
        # Determine date range
        try:
            today = timezone.localdate()
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date() if from_date_str else today
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date() if to_date_str else today
        except (ValueError, TypeError):
            from_date = to_date = today

        # 1. Get all active employees for the company (excluding admin/master roles)
        employee_qs = Employee.objects.filter(company=company, is_active=True).exclude(user__role__in=['admin', 'master']).select_related('department', 'shift_assigned')
        if search:
            employee_qs = employee_qs.filter(
                Q(employee_id__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search)
            )

        # 2. Bulk fetch existing data for the range
        attendance_records = Attendance.objects.filter(
            company=company, 
            date__range=(from_date, to_date)
        ).select_related('employee')
        
        leaves = EmpLeave.objects.filter(
            company=company,
            status='Approved',
            from_date__lte=to_date,
            to_date__gte=from_date
        ).select_related('employee', 'leave_type')
        
        holidays = CalendarEvent.objects.filter(
            company=company,
            is_holiday=True,
            date__range=(from_date, to_date)
        )

        # Map data for fast lookup: { (emp_id, date): object }
        att_map = { (att.employee_id, att.date): att for att in attendance_records }
        holiday_map = { h.date: h.name for h in holidays }

        # Generate the combined list
        results = []
        curr_date = from_date
        while curr_date <= to_date:
            curr_holiday = holiday_map.get(curr_date)
            
            for emp in employee_qs:
                att = att_map.get((emp.id, curr_date))
                
                # Default synthetic row
                row = {
                    'id': att.id if att else None,
                    'employee_id': emp.employee_id,
                    'employee_name': f"{emp.first_name} {emp.last_name}",
                    'date': str(curr_date),
                    'check_in': att.check_in if att else None,
                    'check_out': att.check_out if att else None,
                    'total_work_duration': att.total_work_duration if att else '--',
                    'total_break_time': att.total_break_time if att else '--',
                    'overtime_duration': att.overtime_duration if att else '--',
                    'is_present': att.is_present if att else False,
                    'leave': att.leave_id if att else None,
                    'status': '',
                    'is_late': False
                }

                # Status logic
                if att:
                    if att.leave_id:
                        row['status'] = 'leave'
                    elif not att.is_present and not att.check_in:
                        row['status'] = 'absent'
                    else:
                        row['status'] = 'present'
                    # Call serializer helper or just use data
                else:
                    # Check for Leave
                    matching_leave = next((l for l in leaves if l.employee_id == emp.id and l.from_date <= curr_date <= l.to_date), None)
                    if matching_leave:
                        row['status'] = 'leave'
                        row['leave'] = matching_leave.id
                    elif curr_holiday:
                        row['status'] = 'holiday'
                    else:
                        row['status'] = 'absent'

                # Filtering by calculated status
                if status_filter != 'all' and status_filter != 'all statuses':
                    if row['status'].lower() != status_filter:
                        continue
                
                results.append(row)
            
            curr_date += timedelta(days=1)

        # 3. Pagination
        total_count = len(results)
        page = self.paginate_queryset(results)
        if page is not None:
            # We need to manually serialize since these aren't all model instances
            # But the 'results' are already dicts matching our needs
            return self.get_paginated_response(page)

        return Response(results)

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Provide aggregated statistics for the current filter/date range, including absent employees."""
        # Use the same logic as list but just count
        company = request.user.company
        from_date_str = request.query_params.get('from_date')
        to_date_str = request.query_params.get('to_date')
        search = request.query_params.get('search', '').strip()

        try:
            today = timezone.localdate()
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date() if from_date_str else today
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date() if to_date_str else today
        except (ValueError, TypeError):
            from_date = to_date = today

        employee_qs = Employee.objects.filter(company=company, is_active=True).exclude(user__role__in=['admin', 'master'])
        if search:
            employee_qs = employee_qs.filter(
                Q(employee_id__icontains=search) | Q(first_name__icontains=search) | Q(last_name__icontains=search)
            )
            
        emp_ids = employee_qs.values_list('id', flat=True)
        num_emps = len(emp_ids)
        num_days = (to_date - from_date).days + 1
        total_slots = num_emps * num_days

        # Counts
        present_count = Attendance.objects.filter(
            employee_id__in=emp_ids,
            date__range=(from_date, to_date),
            leave__isnull=True
        ).filter(Q(is_present=True) | Q(check_in__isnull=False)).distinct().count()

        leave_count = EmpLeave.objects.filter(
            employee_id__in=emp_ids,
            status='Approved',
            from_date__lte=to_date,
            to_date__gte=from_date
        ).count()

        # Simplified for responsiveness:
        # Total = all active slots. Absent = Total - (Present + Leave)
        # Note: In a real system, we'd iterate days for leaves.
        absent_count = max(0, total_slots - present_count - leave_count)

        return Response({
            'total': total_slots,
            'present': present_count,
            'absent': absent_count,
            'leave': leave_count,
        })

    def _is_employee_late(self, attendance):
        """Check if an employee was late for their shift"""
        check_in = attendance.check_in
        shift = getattr(attendance.employee, 'shift_assigned', None)

        if not (check_in and shift and shift.checkin):
            return False

        # Convert check-in to local time
        check_in = timezone.localtime(check_in)

        # Construct shift start datetime in local timezone
        shift_start_dt = datetime.combine(check_in.date(), shift.checkin)
        shift_start_dt = timezone.make_aware(shift_start_dt, timezone.get_current_timezone())

        # Add grace period
        allowed_latest_checkin = shift_start_dt + shift.grace()

        return check_in > allowed_latest_checkin

    @action(detail=False, methods=['get'])
    def log(self, request):
        current_date = timezone.localdate()
        month = int(request.query_params.get('month', current_date.month))
        year = int(request.query_params.get('year', current_date.year))

        num_days = monthrange(year, month)[1]
        month_dates = [datetime(year, month, day).date() for day in range(1, num_days + 1)]

        company = request.user.company
        employees = Employee.objects.filter(company=company).exclude(user__role__in=['admin', 'master'])

        attendance_qs = Attendance.objects.filter(
            date__year=year,
            date__month=month,
            company=company
        ).select_related('employee')

        approved_leaves = EmpLeave.objects.filter(
            status='Approved',
            from_date__lte=month_dates[-1],
            to_date__gte=month_dates[0],
            company=company
        ).select_related('leave_type', 'employee')

        holidays = CalendarEvent.objects.filter(
            is_holiday=True,
            date__year=year,
            date__month=month
        )

        def build_valid_days(week_start, week_end):
            DAYS = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
            start_idx = DAYS.index(week_start)
            end_idx = DAYS.index(week_end)
            if start_idx <= end_idx:
                return DAYS[start_idx:end_idx + 1]
            return DAYS[start_idx:] + DAYS[:end_idx + 1]

        attendance_records = {}

        for emp in employees:
            emp_id = emp.id
            dw = DepartmentWiseWorkingDays.objects.filter(department=emp.department).first() if emp.department else None
            valid_days = build_valid_days(dw.week_start_day, dw.week_end_day) if dw else []

            daily_records = {}
            for date in month_dates:
                day_name = date.strftime('%A').lower()
                if valid_days and day_name not in valid_days:
                    continue

                daily_records[str(date)] = {
                    'status': '-',
                    'punch_in': None,
                    'punch_out': None,
                    'worked_hours': 0.0,
                    'is_late': False,
                    'leave_type': '',
                    'is_holiday': False,
                    'holiday_name': '',
                }

            attendance_records[emp_id] = {
                'employee_id': emp.id,
                'employee_name': emp.full_name,
                'daily_records': daily_records,
                'total_hours': 0.0,
                'attendance_percentage': 0.0,
            }

        # Mark holidays
        for holiday in holidays:
            for emp_id, record in attendance_records.items():
                if str(holiday.date) in record['daily_records']:
                    daily = record['daily_records'][str(holiday.date)]
                    daily['status'] = 'H'
                    daily['is_holiday'] = True
                    daily['holiday_name'] = holiday.name

        # Mark approved leaves
        for leave in approved_leaves:
            emp_id = leave.employee.id
            leave_days = [leave.from_date + timedelta(days=i) for i in range((leave.to_date - leave.from_date).days + 1)]
            for day in leave_days:
                day_str = str(day)
                if emp_id in attendance_records and day_str in attendance_records[emp_id]['daily_records']:
                    daily = attendance_records[emp_id]['daily_records'][day_str]
                    if daily['status'] != 'H':
                        daily['status'] = 'L'
                        daily['leave_type'] = leave.leave_type.leave_name if leave.leave_type else ''

        # Punch, worked hours
        for record in attendance_qs:
            emp_id = record.employee.id
            day = str(record.date)

            if emp_id not in attendance_records or day not in attendance_records[emp_id]['daily_records']:
                continue

            daily = attendance_records[emp_id]['daily_records'][day]
            daily['punch_in'] = record.check_in
            daily['punch_out'] = record.check_out

            if record.leave:
                daily['status'] = 'L'
                continue

            shift = getattr(record.employee, 'shift_assigned', None)

            if record.check_in and shift:
                shift_start = timezone.make_aware(datetime.combine(record.date, shift.checkin))
                grace = shift.grace()
                if record.check_in > shift_start + grace:
                    daily['is_late'] = True

            if record.check_in and record.check_out:
                pin = timezone.localtime(record.check_in)
                pout = timezone.localtime(record.check_out)
                worked_seconds = (pout - pin).total_seconds()

                breaks = record.breaks.all()
                for b in breaks:
                    if b.start and b.end:
                        worked_seconds -= (b.end - b.start).total_seconds()

                daily['worked_hours'] = max(round(worked_seconds / 3600, 2), 0.0)

        # Compute totals
        for emp_id, emp_data in attendance_records.items():
            total_present = 0.0
            total_hours = 0.0

            for date, daily in emp_data['daily_records'].items():
                if daily['status'] in ['H', 'L']:
                    total_present += 1
                    continue

                worked_hours = daily['worked_hours']

                shift_record = attendance_qs.filter(employee__id=emp_id, date=date).first()
                shift = getattr(shift_record.employee, 'shift_assigned', None) if shift_record else None

                if not shift:
                    continue

                full_day = shift.full_day_hours()
                half_day = shift.half_day_hours()

                if worked_hours >= full_day:
                    daily['status'] = 'P'
                    total_present += 1
                elif worked_hours >= half_day:
                    daily['status'] = 'H'
                    total_present += 0.5
                else:
                    daily['status'] = 'A'

                total_hours += worked_hours

            num_working_days = len(emp_data['daily_records'])
            emp_data['total_hours'] = round(total_hours, 2)
            emp_data['attendance_percentage'] = round((total_present / num_working_days) * 100, 2) if num_working_days else 0.0
            
            # Add shift info for frontend
            shift = emp.shift_assigned
            emp_data['shift_info'] = {
                'full_day': shift.full_day_hours() if shift else 8.0,
                'half_day': shift.half_day_hours() if shift else 4.0,
                'checkin': shift.checkin.strftime('%H:%M') if shift else '09:00',
                'checkout': shift.checkout.strftime('%H:%M') if shift else '18:00',
            }

        return Response({
            'month_dates': month_dates,
            'attendance_records': list(attendance_records.values()),
        })
 
class AttendanceLogView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]
    
    def post(self, request):
        employee_id = request.data.get('employee_id')
        date_str = request.data.get('date')
        check_in = request.data.get('check_in')
        check_out = request.data.get('check_out')
        remarks = request.data.get('remarks', '')
        status_val = request.data.get('status', None)  # Optional: Present/Absent/Leave/Half Day/Holiday

        if not employee_id or not date_str:
            return Response({'error': 'employee_id and date are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            emp = Employee.objects.get(employee_id=employee_id)
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found.'}, status=status.HTTP_404_NOT_FOUND)

        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        except Exception:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

        att, created = Attendance.objects.get_or_create(
                            employee=emp,
                            date=date_obj,
                            company=emp.company,  # Ensure company is set
                            defaults={}
                        )        # Prefill check-in if exists and no check-out
        prefill = {}
        if att.check_in and not att.check_out:
            prefill['check_in'] = att.check_in.strftime('%H:%M')

        # Update fields if provided
        if status_val:
            if status_val.lower() == 'present':
                att.is_present = True
                if check_in:
                    dt = datetime.combine(date_obj, datetime.strptime(check_in, '%H:%M').time())
                    att.check_in = timezone.make_aware(dt)
                else:
                    att.check_in = None
                    
                if check_out:
                    dt = datetime.combine(date_obj, datetime.strptime(check_out, '%H:%M').time())
                    att.check_out = timezone.make_aware(dt)
                else:
                    att.check_out = None
                
                # Recalculate duration if we have times
                att.calculate_work_duration()
            elif status_val.lower() == 'absent':
                att.is_present = False
                att.check_in = None
                att.check_out = None
                att.total_work_duration = None
                att.overtime_duration = None
                att.total_break_time = None
        else:
            # Default behavior if status not provided: update what's sent
            if check_in is not None:
                if check_in:
                    dt = datetime.combine(date_obj, datetime.strptime(check_in, '%H:%M').time())
                    att.check_in = timezone.make_aware(dt)
                else:
                    att.check_in = None
                    
            if check_out is not None:
                if check_out:
                    dt = datetime.combine(date_obj, datetime.strptime(check_out, '%H:%M').time())
                    att.check_out = timezone.make_aware(dt)
                else:
                    att.check_out = None
            
            att.calculate_work_duration()

        if remarks is not None:
            att.remarks = remarks

        att.save()

        return Response({
            'message': 'Attendance updated successfully.',
            'employee_id': emp.employee_id,
            'date': date_str,
            'is_present': att.is_present,
            'check_in': att.check_in.strftime('%H:%M') if att.check_in else None,
            'check_out': att.check_out.strftime('%H:%M') if att.check_out else None,
            'remarks': att.remarks,
        }, status=status.HTTP_200_OK)
        
    def get(self, request):
        month = request.query_params.get('month')  
        if not month:
            return Response({"error": "Month parameter required (format: YYYY-MM)"}, status=400)

        try:
            year, month_num = map(int, month.split('-'))
            start_date = datetime(year, month_num, 1).date()
            month_end_date = datetime(year, month_num, monthrange(year, month_num)[1]).date()
            today = timezone.localdate()
            # For attendance calculation, only consider up to today if viewing current/future month
            effective_end_date = min(month_end_date, today)
            end_date = month_end_date  # Still use full month for holidays/reference
        except (ValueError, IndexError):
            return Response({"error": "Invalid month format. Use YYYY-MM"}, status=400)

        # ── Backend search & filter on Employee queryset ──
        search_query = request.query_params.get('search', '').strip()
        department_filter = request.query_params.get('department', '').strip()
        status_filter = request.query_params.get('status', '').strip()  # Present/Absent/Leave/Half Day

        # Get holidays for the month
        holidays = CalendarEvent.objects.filter(
            date__range=(start_date, end_date), 
            is_holiday=True,
            company=request.user.company
        )
        holidays_dict = {h.date: h.name for h in holidays}

        # Get employees with related data (excluding admin/master roles)
        employees = Employee.objects.filter(
            company=request.user.company,
            is_active=True
        ).exclude(user__role__in=['admin', 'master']).select_related('department').prefetch_related('attendances')

        # Apply search filter on employees
        if search_query:
            employees = employees.filter(
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(employee_id__icontains=search_query) |
                Q(email__icontains=search_query)
            )

        # Apply department filter
        if department_filter:
            employees = employees.filter(department__department_name__iexact=department_filter)

        employees = employees.order_by('first_name', 'last_name')

        result = []
        
        for emp in employees:
            # Get all attendance records for this employee in the month
            attendance_qs = Attendance.objects.filter(
                employee=emp, 
                date__range=(start_date, end_date)
            ).select_related('employee__shift_assigned').prefetch_related('break_logs')
            
            # Get department working days configuration
            dept_working_days = DepartmentWiseWorkingDays.objects.filter(
                department=emp.department,
                company=request.user.company
            ).first()

            valid_weekday_names = self._valid_weekday_names(dept_working_days)
            
            # Determine working days for this employee (full month for reference)
            all_working_days = self._get_working_days_for_month(
                start_date, end_date, dept_working_days, holidays_dict
            )
            # Working days only up to today (for absent marking & percentage calculation)
            elapsed_working_days = self._get_working_days_for_month(
                start_date, effective_end_date, dept_working_days, holidays_dict
            )
            
            daily_data = []
            present_days = absent_days = leave_days = half_days = late_days = 0
            total_worked_hours = 0.0
            leave_summary = {}

            # Process each attendance record (only days that are working days for this department)
            for att in attendance_qs:
                ad = att.date
                if ad in holidays_dict:
                    continue
                if ad.strftime('%A').lower() not in valid_weekday_names:
                    continue
                daily_record = self._process_attendance_record(att, emp.company)
                daily_data.append(daily_record)
                
                # Count status types
                att_status = daily_record["status"]
                worked_hours = daily_record["worked_hours"]
                
                if att_status == "Present":
                    present_days += 1
                    total_worked_hours += worked_hours
                    if daily_record["is_late"]:
                        late_days += 1
                elif att_status == "Half Day":
                    half_days += 1
                    present_days += 0.5
                    total_worked_hours += worked_hours
                    if daily_record["is_late"]:
                        late_days += 1
                elif att_status == "Leave":
                    leave_days += 1
                    leave_type = daily_record["leave_type"]
                    if leave_type:
                        leave_summary[leave_type] = leave_summary.get(leave_type, 0) + 1
                elif att_status == "Absent":
                    absent_days += 1
                # Note: Holidays are not counted in any category as they're non-working days

            # Add holidays to daily data
            for holiday_date, holiday_name in holidays_dict.items():
                if start_date <= holiday_date <= end_date:
                    daily_data.append({
                        "date": str(holiday_date),
                        "status": "Holiday",
                        "check_in": None,
                        "check_out": None,
                        "worked_hours": 0.0,
                        "scheduled_hours": 0.0,
                        "break_time": 0.0,
                        "overtime_hours": 0.0,
                        "is_late": False,
                        "late_by_minutes": 0,
                        "early_departure": False,
                        "early_departure_minutes": 0,
                        "leave_type": None,
                        "leave_type_initials": None,
                        "half_day": False,
                        "remarks": holiday_name,
                        "shift_type": None
                    })

            # Fill missing days as Absent (only for elapsed working days, not future)
            all_dates = {att.date for att in attendance_qs}
            all_dates.update(holidays_dict.keys())
            
            for single_date in elapsed_working_days:
                if single_date not in all_dates:
                    daily_data.append({
                        "date": str(single_date),
                        "status": "Absent",
                        "check_in": None,
                        "check_out": None,
                        "worked_hours": 0.0,
                        "scheduled_hours": 8.0,  # Default
                        "break_time": 0.0,
                        "overtime_hours": 0.0,
                        "is_late": False,
                        "late_by_minutes": 0,
                        "early_departure": False,
                        "early_departure_minutes": 0,
                        "leave_type": None,
                        "leave_type_initials": None,
                        "half_day": False,
                        "remarks": "No attendance record",
                        "shift_type": None
                    })
                    absent_days += 1

            # Add future working days as "Upcoming" (no absent marking)
            for single_date in all_working_days:
                if single_date > effective_end_date and single_date not in all_dates:
                    daily_data.append({
                        "date": str(single_date),
                        "status": "Upcoming",
                        "check_in": None,
                        "check_out": None,
                        "worked_hours": 0.0,
                        "scheduled_hours": 0.0,
                        "break_time": 0.0,
                        "overtime_hours": 0.0,
                        "is_late": False,
                        "late_by_minutes": 0,
                        "early_departure": False,
                        "early_departure_minutes": 0,
                        "leave_type": None,
                        "leave_type_initials": None,
                        "half_day": False,
                        "remarks": "",
                        "shift_type": None
                    })

            # Department off-days (weekdays not in working_days / weekend_days) show as Holiday (★)
            by_date = {d["date"]: i for i, d in enumerate(daily_data)}
            wd_list = (dept_working_days.weekend_days or []) if dept_working_days else []
            off_cursor = start_date
            while off_cursor <= end_date:
                if off_cursor in holidays_dict:
                    off_cursor += timedelta(days=1)
                    continue
                day_key = off_cursor.strftime('%A').lower()
                if day_key in valid_weekday_names:
                    off_cursor += timedelta(days=1)
                    continue
                iso = str(off_cursor)
                remark = self._off_day_remark(off_cursor, wd_list)
                rec = {
                    "date": iso,
                    "status": "Holiday",
                    "check_in": None,
                    "check_out": None,
                    "worked_hours": 0.0,
                    "scheduled_hours": 0.0,
                    "break_time": 0.0,
                    "overtime_hours": 0.0,
                    "is_late": False,
                    "late_by_minutes": 0,
                    "early_departure": False,
                    "early_departure_minutes": 0,
                    "leave_type": None,
                    "leave_type_initials": None,
                    "half_day": False,
                    "remarks": remark,
                    "shift_type": None
                }
                if iso in by_date:
                    daily_data[by_date[iso]] = rec
                else:
                    daily_data.append(rec)
                    by_date[iso] = len(daily_data) - 1
                off_cursor += timedelta(days=1)

            daily_data.sort(key=lambda x: x["date"])

            # Calculate totals and percentages
            total_working_days = len(all_working_days)
            elapsed_working_day_count = len(elapsed_working_days)
            total_days_present = present_days  # This includes half days as 0.5
            
            # Attendance percentage based on elapsed working days only (not future)
            attendance_percentage = (total_days_present / elapsed_working_day_count * 100) if elapsed_working_day_count > 0 else 0
            
            # Average working hours per present day
            avg_hours_per_day = total_worked_hours / present_days if present_days > 0 else 0
            
            # Calculate monthly expected working hours
            total_expected_hours = 0.0
            total_overtime_hours = 0.0
            total_break_time = 0.0
            
            for daily_record in daily_data:
                if daily_record["status"] not in ["Holiday"]:
                    total_expected_hours += daily_record["scheduled_hours"]
                    total_overtime_hours += daily_record["overtime_hours"]
                    total_break_time += daily_record["break_time"]
            
            # Working hours efficiency (actual vs expected)
            hours_efficiency = (total_worked_hours / total_expected_hours * 100) if total_expected_hours > 0 else 0
            
            # Working hours shortage/surplus
            hours_variance = total_worked_hours - total_expected_hours

            # Get company's shift policies for reference
            company_shifts = ShiftPolicy.objects.filter(company=request.user.company)
            shift_policies_info = [
                {
                    "id": shift.id,
                    "shift_type": shift.shift_type or f"Shift {shift.id}",
                    "full_day_hours": shift.full_day_hours(),
                    "half_day_hours": shift.half_day_hours(),
                    "checkin": shift.checkin.strftime('%H:%M') if shift.checkin else None,
                    "checkout": shift.checkout.strftime('%H:%M') if shift.checkout else None,
                    "grace_period_minutes": int(shift.grace().total_seconds() / 60) if shift.grace() else 0
                }
                for shift in company_shifts
            ]

            # Determine today's status for this employee (for status filter)
            today_iso = str(today)
            today_record = next((d for d in daily_data if d["date"] == today_iso), None)
            today_status = today_record["status"] if today_record else "Absent"

            result.append({
                "employee_id": emp.employee_id,
                "employee_name": emp.full_name,
                "department": emp.department.department_name if emp.department else None,
                "month": month,
                "today_status": today_status,
                
                # Monthly Working Days Statistics
                "total_working_days": total_working_days,
                "total_present_days": round(total_days_present, 2),
                "total_absent_days": absent_days,
                "total_leave_days": leave_days,
                "total_half_days": half_days,
                "total_late_days": late_days,
                "total_holidays": len(holidays_dict),
                
                # Monthly Working Hours Statistics
                "total_worked_hours": round(total_worked_hours, 2),
                "total_expected_hours": round(total_expected_hours, 2),
                "total_overtime_hours": round(total_overtime_hours, 2),
                "total_break_time": round(total_break_time, 2),
                "hours_variance": round(hours_variance, 2),  # Positive = overtime, Negative = shortage
                
                # Monthly Percentages & Averages
                "percentage_present": round(attendance_percentage, 2),
                "hours_efficiency": round(hours_efficiency, 2),  # Actual vs Expected hours percentage
                "average_hours_per_day": round(avg_hours_per_day, 2),
                "average_hours_per_working_day": round(total_worked_hours / total_working_days, 2) if total_working_days > 0 else 0,
                
                # Additional Monthly Insights
                "monthly_summary": {
                    "productive_days": present_days + half_days,  # Days with any work done
                    "non_productive_days": absent_days,
                    "leave_utilization": leave_days,
                    "punctuality_score": round((present_days - late_days) / present_days * 100, 2) if present_days > 0 else 100,
                    "overtime_frequency": sum(1 for d in daily_data if d["overtime_hours"] > 0),
                    "break_usage_hours": round(total_break_time, 2)
                },
                
                # Reference Data
                "holidays": [{"date": str(d), "name": n} for d, n in holidays_dict.items()],
                "leave_summary": leave_summary,
                "shift_policies": shift_policies_info,
                "daily_attendance": sorted(daily_data, key=lambda x: x["date"])
            })

        # ── Apply today's status filter after building result ──
        if status_filter and status_filter.lower() != 'all':
            result = [r for r in result if r.get("today_status", "").lower() == status_filter.lower()]

        # ── Pagination ──
        total_count = len(result)
        try:
            page = int(request.query_params.get('page', 1))
        except (ValueError, TypeError):
            page = 1
        try:
            page_size = int(request.query_params.get('page_size', 10))
        except (ValueError, TypeError):
            page_size = 10
        page_size = min(page_size, 100)  # Max 100

        total_pages = max(1, -(-total_count // page_size))  # ceil division
        page = max(1, min(page, total_pages))
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        paginated_result = result[start_idx:end_idx]

        # Collect unique departments for filter dropdown
        all_departments = list(
            Employee.objects.filter(
                company=request.user.company,
                is_active=True,
                department__isnull=False
            ).values_list('department__department_name', flat=True).distinct().order_by('department__department_name')
        )

        return Response({
            "count": total_count,
            "page": page,
            "page_size": page_size,
            "total_pages": total_pages,
            "departments": all_departments,
            "results": paginated_result
        })

    def _valid_weekday_names(self, dept_working_days):
        """Lowercase weekday names that count as working days for the department."""
        weekdays = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
        if dept_working_days:
            wd = dept_working_days.working_days
            if isinstance(wd, list) and len(wd) > 0:
                return [str(x).lower() for x in wd]
            ws = (dept_working_days.week_start_day or '').strip().lower()
            we = (dept_working_days.week_end_day or '').strip().lower()
            if ws in weekdays and we in weekdays:
                si, ei = weekdays.index(ws), weekdays.index(we)
                if si <= ei:
                    return weekdays[si : ei + 1]
                return weekdays[si:] + weekdays[: ei + 1]
        # No department row: match prior behavior (Mon–Sat as working; Sunday off)
        return ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday']

    def _off_day_remark(self, d, weekend_days_list):
        """Tooltip text for a non-working weekday (dept schedule)."""
        long_name = d.strftime('%A')
        if isinstance(weekend_days_list, list):
            for entry in weekend_days_list:
                if str(entry).lower() == long_name.lower():
                    return f'Off day ({entry})'
        return 'Off day (non-working)'

    def _get_working_days_for_month(self, start_date, end_date, dept_working_days, holidays):
        """Dates in range that are working days for the department (excludes company holidays)."""
        working_days = []
        current_date = start_date
        valid_weekdays = self._valid_weekday_names(dept_working_days)

        while current_date <= end_date:
            day_name = current_date.strftime('%A').lower()
            if day_name in valid_weekdays and current_date not in holidays:
                working_days.append(current_date)
            current_date += timedelta(days=1)

        return working_days

    def _process_attendance_record(self, attendance, company):
        """Process a single attendance record and return comprehensive data"""
        # shift_policy = attendance.shift
        shift_policy = attendance.employee.shift_assigned  

        
        # If no shift assigned, get company's default shift
        if not shift_policy:
            shift_policy = ShiftPolicy.objects.filter(company=company).first()
        
        # Calculate basic worked hours
        worked_hours = 0.0
        scheduled_hours = shift_policy.full_day_hours() if shift_policy else 8.0
        break_time = 0.0
        overtime_hours = 0.0
        
        if attendance.check_in and attendance.check_out:
            # Calculate total time between check-in and check-out
            check_in_dt = attendance.check_in
            check_out_dt = attendance.check_out
            
            # Convert to local time for calculation
            check_in_local = localtime(check_in_dt)
            check_out_local = localtime(check_out_dt)
            
            total_seconds = (check_out_local - check_in_local).total_seconds()
            
            # Deduct break time
            break_logs = attendance.break_logs.all()
            total_break_seconds = 0
            
            for break_log in break_logs:
                if break_log.start and break_log.end:
                    break_duration = (break_log.end - break_log.start).total_seconds()
                    total_break_seconds += break_duration
            
            break_time = round(total_break_seconds / 3600, 2)
            worked_hours = round((total_seconds - total_break_seconds) / 3600, 2)
            worked_hours = max(0, worked_hours)  # Ensure non-negative
            
            # Calculate overtime (hours worked beyond scheduled hours)
            if worked_hours > scheduled_hours:
                overtime_hours = worked_hours - scheduled_hours

        # Determine status based on worked hours and shift policy
        status = "Absent"
        half_day = False
        
        if attendance.leave:
            status = "Leave"
        elif worked_hours > 0 and shift_policy:
            if worked_hours >= shift_policy.full_day_hours():
                status = "Present"
            elif worked_hours >= shift_policy.half_day_hours():
                status = "Half Day"
                half_day = True
            else:
                status = "Absent"  # Insufficient hours worked
        elif worked_hours > 0:
            # No shift policy, use default thresholds
            if worked_hours >= 8.0:
                status = "Present"
            elif worked_hours >= 4.0:
                status = "Half Day"
                half_day = True
            else:
                status = "Absent"

        # Check if late
        is_late = False
        late_minutes = 0
        if shift_policy and attendance.check_in:
            scheduled_checkin = datetime.combine(attendance.date, shift_policy.checkin)
            actual_checkin = datetime.combine(attendance.date, localtime(attendance.check_in).time())
            grace_period = shift_policy.grace()
            
            if actual_checkin > (scheduled_checkin + grace_period):
                is_late = True
                late_minutes = int((actual_checkin - scheduled_checkin).total_seconds() / 60)

        # Check for early departure
        early_departure = False
        early_departure_minutes = 0
        if shift_policy and attendance.check_out:
            scheduled_checkout = datetime.combine(attendance.date, shift_policy.checkout)
            actual_checkout = datetime.combine(attendance.date, localtime(attendance.check_out).time())
            
            if actual_checkout < scheduled_checkout:
                early_departure = True
                early_departure_minutes = int((scheduled_checkout - actual_checkout).total_seconds() / 60)

        # Get leave information
        leave_type_val = None
        leave_type_initials = None
        if attendance.leave and attendance.leave.leave_type:
            leave_type_val = attendance.leave.leave_type.leave_name
            leave_type_initials = leave_type_val[:2].upper() if leave_type_val else None

        return {
            "date": str(attendance.date),
            "status": status,
            "check_in": localtime(attendance.check_in).strftime("%H:%M") if attendance.check_in else None,
            "check_out": localtime(attendance.check_out).strftime("%H:%M") if attendance.check_out else None,
            "worked_hours": worked_hours,
            "scheduled_hours": scheduled_hours,
            "break_time": break_time,
            "overtime_hours": overtime_hours,
            "is_late": is_late,
            "late_by_minutes": late_minutes,
            "early_departure": early_departure,
            "early_departure_minutes": early_departure_minutes,
            "leave_type": leave_type_val,
            "leave_type_initials": leave_type_initials,
            "half_day": half_day,
            "remarks": attendance.remarks or "",
            "shift_type": shift_policy.shift_type if shift_policy else None
        }


class CompanyPoliciesViewSet(viewsets.ModelViewSet):
    queryset = CompanyPolicies.objects.all()
    serializer_class = PolicyConfigurationSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get_queryset(self):
        return CompanyPolicies.objects.filter(company=self.request.user.company)

    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company)
 
class ApprovedLeaveLogView(generics.ListAPIView):
    """
    Admin approved leave logs with backend pagination + search.
    """
    serializer_class = LeaveLogSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = [
        "employee__first_name",
        "employee__last_name",
        "employee__email",
        "reporting_manager__first_name",
        "reporting_manager__last_name",
        "reason",
        "leave_type__leave_name",
    ]

    def get_queryset(self):
        qs = EmpLeave.objects.filter(
            status="Approved",
            company=self.request.user.company,
        ).select_related("employee", "reporting_manager", "leave_type")

        employee_id = self.request.query_params.get("employee_id")
        from_date = self.request.query_params.get("from_date")
        to_date = self.request.query_params.get("to_date")

        if employee_id:
            qs = qs.filter(employee__id=employee_id)
        if from_date:
            qs = qs.filter(from_date__gte=from_date)
        if to_date:
            qs = qs.filter(from_date__lte=to_date)
        return qs.order_by("-from_date", "-id")


class RejectedLeaveLogView(generics.ListAPIView):
    serializer_class = LeaveLogSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = [
        "employee__first_name",
        "employee__last_name",
        "employee__email",
        "reporting_manager__first_name",
        "reporting_manager__last_name",
        "reason",
        "leave_type__leave_name",
    ]

    def get_queryset(self):
        qs = EmpLeave.objects.filter(
            status='Rejected',
            company=self.request.user.company
        ).select_related('employee', 'reporting_manager', 'leave_type')

        from_date = self.request.query_params.get('from_date')
        to_date = self.request.query_params.get('to_date')
        employee_id = self.request.query_params.get('employee_id')

        if from_date:
            qs = qs.filter(from_date__gte=from_date)
        if to_date:
            qs = qs.filter(from_date__lte=to_date)
        if employee_id:
            qs = qs.filter(employee__id=employee_id)

        return qs.order_by("-from_date", "-id")

class PendingLeaveLogView(generics.ListAPIView):
    """
    Admin pending leave requests with backend pagination + search.
    """
    serializer_class = LeaveLogSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = [
        "employee__first_name",
        "employee__last_name",
        "employee__employee_id",
        "employee__email",
        "reporting_manager__first_name",
        "reporting_manager__last_name",
        "reason",
        "leave_type__leave_name",
    ]

    def get_queryset(self):
        qs = EmpLeave.objects.filter(
            status="Pending",
            company=self.request.user.company,
        ).select_related("employee", "reporting_manager", "leave_type")

        from_date = self.request.query_params.get("from_date")
        to_date = self.request.query_params.get("to_date")

        if from_date:
            qs = qs.filter(from_date__gte=from_date)
        if to_date:
            qs = qs.filter(from_date__lte=to_date)
        return qs.order_by("-from_date", "-id")

class LeaveHistoryView(generics.ListAPIView):
    """
    Admin combined leave history (Approved/Rejected) with backend pagination + search.
    """
    serializer_class = LeaveLogSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = [
        "employee__first_name",
        "employee__last_name",
        "employee__employee_id",
        "employee__email",
        "reporting_manager__first_name",
        "reporting_manager__last_name",
        "reason",
        "leave_type__leave_name",
    ]

    def get_queryset(self):
        # Default to Approved/Rejected, but allow filtering by status if provided
        status_filter = self.request.query_params.get("status", "history")
        
        qs = EmpLeave.objects.filter(company=self.request.user.company).select_related("employee", "reporting_manager", "leave_type")
        
        if status_filter == "history":
            qs = qs.filter(status__in=["Approved", "Rejected"])
        elif status_filter != "all":
            qs = qs.filter(status=status_filter)

        from_date = self.request.query_params.get("from_date")
        to_date = self.request.query_params.get("to_date")

        if from_date:
            qs = qs.filter(from_date__gte=from_date)
        if to_date:
            qs = qs.filter(from_date__lte=to_date)

        return qs.order_by("-from_date", "-id")

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            response = self.get_paginated_response(serializer.data)
            
            # Add extra stats for the dashboard cards
            stats_qs = EmpLeave.objects.filter(company=request.user.company)
            response.data['approved_count'] = stats_qs.filter(status='Approved').count()
            response.data['rejected_count'] = stats_qs.filter(status='Rejected').count()
            
            return response

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

class AdminApproveEmpLeaveView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def post(self, request, leave_id):
        company = request.user.company
        leave = get_object_or_404(EmpLeave, id=leave_id, company=company)
        if leave.status != 'Approved':
            leave.status = 'Approved'
            leave.save()
            return Response({'detail': 'Leave approved successfully.'})
        return Response({'detail': 'This leave is already approved.'}, status=400)

class AdminRejectEmpLeaveView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def post(self, request, leave_id):
        company = request.user.company
        leave = get_object_or_404(EmpLeave, id=leave_id, company=company)
        if leave.status != 'Rejected':
            rejection_reason = request.data.get('reason', '')
            leave.status = 'Rejected'
            leave.rejection_reason = rejection_reason
            leave.save()
            return Response({'detail': 'Leave rejected successfully.'})
        return Response({'detail': 'This leave is already rejected.'}, status=400)
   
class UserLogListView(generics.ListAPIView):
    serializer_class = UserLogSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get_queryset(self):
        qs = UserRegister.objects.filter(company=self.request.user.company)
        username = self.request.query_params.get('username')
        if username:
            qs = qs.filter(username=username)
        return qs


class UserLogDeleteView(generics.DestroyAPIView):
    serializer_class = UserLogSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get_queryset(self):
        return UserRegister.objects.filter(company=self.request.user.company)


class BreakConfigViewSet(viewsets.ModelViewSet):
    serializer_class = BreakConfigSerializer
    permission_classes = [IsAuthenticated,IsAdminUser]

    def get_queryset(self):
        return BreakConfig.objects.filter(company=self.request.user.company)

    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company)
        
        
        
class LetterTemplateViewSet(viewsets.ModelViewSet):
    serializer_class = LetterTemplateSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['title', 'content', 'email_content']

    def get_queryset(self):
        return LetterTemplate.objects.filter(
            company=self.request.user.company
        ).order_by("-created_at")

    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company, created_by=self.request.user)

    @action(detail=True, methods=["post"], url_path="preview")
    def preview(self, request, pk=None):
        """
        Preview a letter template with candidate data before sending
        """
        template = self.get_object()
        candidate_data = request.data.get("candidate_data", {})

        try:
            # Use safe Python string.Template
            tmpl = string.Template(template.content)
            rendered_text = tmpl.safe_substitute(candidate_data)  
            
            return Response({
                "template_id": template.id,
                "preview_text": rendered_text,
                "input_data": candidate_data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        


class GenerateLetterContentAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]


    def post(self, request):
        
        template_id = request.data.get('template_id')
        letter_type = request.data.get('type')
        employee_id = request.data.get('employee_id')
        candidate_id = request.data.get('candidate_id')
        relieved_employee_id = request.data.get('relieved_employee_id')
        email_content = request.data.get('email_content')

        if not letter_type:
            return Response({'error': 'Letter type is required.'}, status=status.HTTP_400_BAD_REQUEST)

        # Fetch template
        try:
            template = LetterTemplate.objects.get(id=template_id, company=request.user.company)
        except LetterTemplate.DoesNotExist:
            return Response({'error': 'Template not found'}, status=status.HTTP_404_NOT_FOUND)

        # Determine which model to use
        obj = None
        obj_type = None
        recipient_email = None
        if employee_id:
            try:
                obj = Employee.objects.select_related('company', 'department', 'designation').get(id=employee_id, company=request.user.company)
                obj_type = 'employee'
                recipient_email = obj.email
            except Employee.DoesNotExist:
                return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)
        elif candidate_id:
            try:
                obj = Recruitment.objects.get(id=candidate_id)
                obj_type = 'candidate'
                recipient_email = obj.email
            except Recruitment.DoesNotExist:
                return Response({'error': 'Candidate not found'}, status=status.HTTP_404_NOT_FOUND)
        elif relieved_employee_id:
            try:
                obj = RelievedEmployee.objects.select_related('employee').get(id=relieved_employee_id)
                obj_type = 'relievedemployee'
                recipient_email = obj.employee.email if obj.employee else None
            except RelievedEmployee.DoesNotExist:
                return Response({'error': 'Relieved employee not found'}, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({'error': 'No valid person id provided.'}, status=status.HTTP_400_BAD_REQUEST)

        # Build data dict for placeholders
        data = {}
        if obj_type == 'employee':
            data = {
                'name': obj.full_name,
                'employee_id': obj.employee_id,
                'designation': obj.designation.designation_name if obj.designation else '',
                'department': obj.department.department_name if obj.department else '',
                'joining_date': obj.date_of_joining.strftime('%Y-%m-%d') if obj.date_of_joining else '',
                'last_working_date': obj.date_of_releaving.strftime('%Y-%m-%d') if obj.date_of_releaving else '',
                'ctc': str(obj.ctc) if obj.ctc else '',
                'company': obj.company.name if obj.company else '',
                'location': obj.company.location if obj.company else '',
            }
        elif obj_type == 'candidate':
            data = {
                'name': obj.name,
                'designation': obj.job_title,
                'joining_date': obj.appointment_date.strftime('%Y-%m-%d') if obj.appointment_date else '',
                'ctc': str(obj.salary) if obj.salary else '',
                'company': request.user.company.name if hasattr(request.user, 'company') and request.user.company else '',
                'location': request.user.company.location if hasattr(request.user, 'company') and request.user.company else '',
                'address': obj.address or '',
            }
        elif obj_type == 'relievedemployee':
            emp = obj.employee
            data = {
                'name': emp.full_name,
                'employee_id': emp.employee_id,
                'designation': emp.designation.designation_name if emp.designation else '',
                'department': emp.department.department_name if emp.department else '',
                'joining_date': emp.date_of_joining.strftime('%Y-%m-%d') if emp.date_of_joining else '',
                'last_working_date': obj.relieving_date.strftime('%Y-%m-%d') if obj.relieving_date else '',
                'ctc': str(emp.ctc) if emp.ctc else '',
                'company': emp.company.name if emp.company else '',
                'location': emp.company.location if emp.company else '',
            }

        # Find all placeholders in the template
        placeholders = set(re.findall(r'<(\w+)>', template.content))

        # Replace placeholders in letter content
        def replacer(match):
            key = match.group(1)
            return str(data.get(key, f'<{key}>'))
        filled_content = re.sub(r'<(\w+)>', replacer, template.content)

        # Replace placeholders in email_content (only <name> and <company> for safety)
        if email_content:
            email_content = re.sub(r'<(name|company)>', lambda m: str(data.get(m.group(1), f'<{m.group(1)}>')), email_content)

        generated_letter = None
        if obj_type == 'candidate':
            generated_letter, created = GeneratedLetter.objects.get_or_create(
                candidate=obj,
                template=template,
                type=letter_type,
                defaults={
                    'content': filled_content,
                    'title': template.title,
                }
            )
            if not created:
                generated_letter.content = filled_content
                generated_letter.title = template.title
                generated_letter.save()
        elif obj_type == 'employee':
            generated_letter, created = GeneratedLetter.objects.get_or_create(
                employee=obj,
                template=template,
                type=letter_type,
                defaults={
                    'content': filled_content,
                    'title': template.title,
                }
            )
            if not created:
                generated_letter.content = filled_content
                generated_letter.title = template.title
                generated_letter.save()
        elif obj_type == 'relievedemployee':
            generated_letter, created = GeneratedLetter.objects.get_or_create(
                relieved_employee=obj,
                template=template,
                type=letter_type,
                defaults={
                    'content': filled_content,
                    'title': template.title,
                }
            )
            if not created:
                generated_letter.content = filled_content
                generated_letter.title = template.title
                generated_letter.save()

        # --- EMAIL SENDING LOGIC ---
        # Only send if not already sent
        if hasattr(generated_letter, 'sent') and generated_letter.sent:
            return Response({
                'content': filled_content,
                'placeholders': list(placeholders),
                'filled': data,
                'generated_letter_id': generated_letter.id if generated_letter else None,
                'email_status': 'already_sent'
            }, status=status.HTTP_200_OK)

        # Generate PDF (implement generate_letter_pdf to return PDF bytes)
        try:
            # Determine company for PDF context
            if obj_type == 'candidate':
                company = request.user.company
            elif obj_type == 'employee':
                company = obj.company
            elif obj_type == 'relievedemployee':
                company = emp.company
            else:
                company = None
            pdf_bytes = generate_letter_pdf(company, template.title, filled_content, request)
        except Exception as e:
            return Response({'error': f'PDF generation failed: {str(e)}'}, status=500)

        # Send email if recipient email is present (always send for offer/appointment letters)
        if recipient_email:
            try:
                # Use email_content if available, otherwise use a default message
                email_body = email_content if email_content else f"Please find attached the {template.title} for your review."
                
                email = EmailMessage(
                    subject=template.title,
                    body=email_body,
                    to=[recipient_email]
                )
                email.attach(f"{template.title}.pdf", pdf_bytes, "application/pdf")
                email.send()
                # Mark as sent
                generated_letter.sent = True
                generated_letter.sent_at = timezone.now()
                generated_letter.save()
                email_status = 'sent'
            except Exception as e:
                email_status = f'error: {str(e)}'
        else:
            email_status = 'no_recipient_email'

        return Response({
            'content': filled_content,
            'placeholders': list(placeholders),
            'filled': data,
            'generated_letter_id': generated_letter.id if generated_letter else None,
            'email_status': email_status
        }, status=status.HTTP_200_OK)
        


class GeneratedLetterViewSet(viewsets.ModelViewSet):
    queryset = GeneratedLetter.objects.all()
    serializer_class = GeneratedLetterSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        letter_type = self.request.query_params.get('type')
        template_id = self.request.query_params.get('template_id')
        candidate_id = self.request.query_params.get('candidate_id')
        relieved_id = self.request.query_params.get('relieved_id')

        if letter_type:
            queryset = queryset.filter(type=letter_type)
        if template_id:
            queryset = queryset.filter(template_id=template_id)
        if candidate_id:
            queryset = queryset.filter(candidate_id=candidate_id)
        if relieved_id:
            queryset = queryset.filter(relieved_employee_id=relieved_id)
        return queryset

    def perform_create(self, serializer):
        instance = serializer.save()
        uploaded_file = self.request.FILES.get('file')
        if uploaded_file:
            # Save the file and set file_path (adjust path as needed)
            instance.file_path = f'letters/{uploaded_file.name}'
            instance.save()


from rest_framework.permissions import IsAuthenticated,AllowAny

class RefreshTokenView(APIView):
    permission_classes = [AllowAny] 

    def post(self, request):
        try:
            serializer = RefreshTokenSerializer(data=request.data)
            if serializer.is_valid():
                return Response(serializer.validated_data)
            return Response({"status":"failed","response_code":status.HTTP_404_NOT_FOUND,"message":serializer.errors})
        except Exception as e:
            message = str(e)
            return Response({"status":"failed","response_code":status.HTTP_500_INTERNAL_SERVER_ERROR,"message":message})
        


class AssignShiftAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        """
        Assign a shift to an employee.
        Expects JSON: { "employee_id": 1, "shift_id": 2 }
        """
        employee_id = request.data.get("employee_id")
        shift_id = request.data.get("shift_id")

        try:
            employee = Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            return Response({"error": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)

        try:
            shift = ShiftPolicy.objects.get(id=shift_id)
        except ShiftPolicy.DoesNotExist:
            return Response({"error": "Shift not found"}, status=status.HTTP_404_NOT_FOUND)

        employee.shift_assigned = shift
        employee.save()

        serializer = AssignShiftSerializer(employee)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
class UserProfileView(generics.RetrieveAPIView):
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        return self.request.user

class UserUpdateView(generics.UpdateAPIView):
    serializer_class = UserUpdateSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_object(self):
        return self.request.user

    def update(self, request, *args, **kwargs):
        user = self.get_object()
        # Use request.data which handles both JSON and Multipart
        serializer = self.get_serializer(data=request.data, context={'request': request})
        try:
            serializer.is_valid(raise_exception=True)
            data = serializer.validated_data
            files = request.FILES

            # Update User fields (Restrict sensitive fields for employees if needed, 
            # but usually first_name/last_name are okay)
            if 'username' in data:
                new_username = data['username']
                if UserRegister.objects.exclude(pk=user.pk).filter(username=new_username).exists():
                    return Response({"detail": "Username already exists."}, status=status.HTTP_400_BAD_REQUEST)
                user.username = new_username
            
            if 'first_name' in data:
                user.first_name = data['first_name']
            if 'last_name' in data:
                user.last_name = data['last_name']
            
            # Update password if provided
            if 'new_password' in data:
                user.set_password(data['new_password'])

            user.save()

            # Update Employee Profile (Ensure it exists for Master/Admin too if they want a photo)
            employee = user.employee_profile
            if not employee:
                # Create profile if missing (needed for photo/mobile/address storage)
                company = user.company or Company.objects.first()
                if company:
                    employee = Employee.objects.create(
                        user=user, 
                        company=company,
                        first_name=user.first_name,
                        last_name=user.last_name
                    )

            if employee:
                if 'first_name' in data:
                    employee.first_name = data['first_name']
                if 'last_name' in data:
                    employee.last_name = data['last_name']
                if 'mobile' in data:
                    employee.mobile = data['mobile']
                if 'address' in data:
                    employee.permanent_address = data['address']
                if 'location' in data:
                    employee.temporary_address = data['location']
                
                # Handle missing details
                if 'aadhar_no' in data:
                    employee.aadhar_no = data['aadhar_no']
                if 'pan_no' in data:
                    employee.pan_no = data['pan_no']
                if 'guardian_name' in data:
                    employee.guardian_name = data['guardian_name']
                if 'guardian_mobile' in data:
                    employee.guardian_mobile = data['guardian_mobile']
                if 'gender' in data:
                    employee.gender = data['gender']
                if 'date_of_birth' in data:
                    employee.date_of_birth = data['date_of_birth']
                if 'bank_name' in data:
                    employee.bank_name = data['bank_name']
                if 'account_no' in data:
                    employee.account_no = data['account_no']
                if 'ifsc_code' in data:
                    employee.ifsc_code = data['ifsc_code']
                if 'payment_method' in data:
                    employee.payment_method = data['payment_method']
                
                # Handle Files
                if 'photo' in files:
                    employee.photo = files['photo']
                if 'aadhar_card' in files:
                    employee.aadhar_card = files['aadhar_card']
                if 'pan_card' in files:
                    employee.pan_card = files['pan_card']
                
                employee.save()

            return Response({"detail": "Profile updated successfully."}, status=status.HTTP_200_OK)
        except ValidationError as e:
            return Response(e.detail, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
import random
from django.contrib.auth.hashers import make_password


class SendOtpView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get("email")
        if not email:
            return Response({"error": "Email is required"})

        otp = str(random.randint(100000, 999999))

        otp_obj, created = EmailOTP.objects.get_or_create(email=email)
        otp_obj.otp = otp
        otp_obj.created_at = timezone.now()  
        otp_obj.save()

        subject = "Your OTP for Password Reset"
        message = f"""
        Hello,

        You have requested to reset your password for your account.

        Your One-Time Password (OTP) is: {otp}

        This OTP is valid for the next 5 minutes. Please do not share this code with anyone for security reasons.

        If you did not request this, please ignore this email or contact support immediately.

        Thank you,  
        Team Innovyx Tech Labs
        """
        from_email = settings.EMAIL_HOST_USER

        try:
            send_mail(subject, message, from_email, [email])
            return Response({"message": "OTP sent successfully"})
        except Exception as e:
            return Response({"error": f"Failed to send email: {str(e)}"})


class VerifyOTPView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get('email')
        otp_input = request.data.get('otp')

        try:
            otp_obj = EmailOTP.objects.get(email=email)

            if otp_obj.is_expired():
                return Response({'error': 'OTP expired'})

            if otp_obj.otp == otp_input:
                otp_obj.verified = True  
                otp_obj.save()
                return Response({'message': 'OTP verified successfully'})
            else:
                return Response({'error': 'Invalid OTP'})

        except EmailOTP.DoesNotExist:
            return Response({'error': 'OTP not found'})

class ResetPasswordView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get("email")
        new_password = request.data.get("new_password")
        confirm_password = request.data.get("confirm_password")

        if not all([email, new_password, confirm_password]):
            return Response({"error": "All fields are required."})

        if new_password != confirm_password:
            return Response({"error": "Passwords do not match."})

        try:
            otp_obj = EmailOTP.objects.get(email=email)
            if not otp_obj.verified:
                return Response({"error": "OTP not verified for this email."})
        except EmailOTP.DoesNotExist:
            return Response({"error": "OTP not found. Please verify OTP first."})

        try:
            user = UserRegister.objects.get(email=email)
            user.password = make_password(new_password)
            user.save()
            otp_obj.verified = False
            otp_obj.save()
            return Response({"message": "Password reset successful."})
        except UserRegister.DoesNotExist:
            return Response({"error": "User not found."})


class EmployeeStatusViewSet(viewsets.ModelViewSet):
    serializer_class = EmployeeStatusSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        # If employee, filter by their company only
        if user.role == 'employee':
            return Employee.objects.filter(company=user.company)
        # Otherwise (admin, master), return all employees
        return Employee.objects.all()
    
class PayrollAttendanceSummaryView(APIView):
    """Return attendance summary for a given employee + date range."""
    permission_classes = [IsAuthenticated, IsAdminUser | IsMaster]

    def get(self, request):
        emp_id = request.query_params.get('employee_id')
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')

        if not emp_id or not from_date or not to_date:
            return Response({'error': 'employee_id, from_date, to_date are required'}, status=400)

        try:
            from datetime import datetime as dt
            from_dt = dt.strptime(from_date, '%Y-%m-%d').date()
            to_dt = dt.strptime(to_date, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

        try:
            employee = Employee.objects.get(id=emp_id)
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=404)

        # 1. Attendance Metrics (Source of Truth)
        m = compute_attendance_metrics(employee, from_dt, to_dt)

        # 2. Leave Details
        leaves = EmpLeave.objects.filter(
            employee=employee, status='Approved',
            from_date__lte=to_dt, to_date__gte=from_dt
        ).select_related('leave_type')
        
        leave_details = []
        for lv in leaves:
            l_start = max(lv.from_date, from_dt)
            l_end = min(lv.to_date, to_dt)
            num_days = (l_end - l_start).days + 1
            leave_details.append({
                'leave_type': lv.leave_type.leave_name if lv.leave_type else 'Unknown',
                'is_paid': lv.leave_type.is_paid if lv.leave_type else False,
                'from_date': str(lv.from_date),
                'to_date': str(lv.to_date),
                'days': num_days,
                'status': lv.status,
            })

        # 3. Reimbursements
        reimbursements = ReimbursementRequest.objects.filter(
            employee=employee, status='approved',
            created_at__date__gte=from_dt, created_at__date__lte=to_dt
        ).select_related('category')
        
        total_reimbursement = sum(r.amount for r in reimbursements)
        reimbursement_details = [{
            'category': r.category.name if r.category else r.custom_category,
            'amount': float(r.amount),
            'date': r.created_at.date().isoformat(),
            'description': r.description
        } for r in reimbursements]

        # 4. Asset Deductions
        asset_reqs = AssetRequest.objects.filter(
            requested_by=employee,
            approval_status='approved',
            updated_at__date__range=[from_dt, to_dt]
        ).select_related('related_fixed_asset', 'related_supply_item').distinct()
        
        asset_deduction_details = []
        supply_batches = {}
        total_asset_deduction = Decimal('0.00')
        
        for r in asset_reqs:
            amt = r.employee_payment_amount or Decimal('0')
            if r.related_fixed_asset:
                asset_name = f"{r.related_fixed_asset.asset_tag} ({r.related_fixed_asset.model_brand or 'No Brand'})"
                asset_deduction_details.append({
                    'id': str(r.id),
                    'item_name': asset_name,
                    'amount': float(amt),
                    'date': r.updated_at.date().isoformat(),
                    'action_type': r.admin_action_type or 'Asset Deduction'
                })
                total_asset_deduction += amt
            elif r.related_supply_item:
                b_id = r.batch_id if r.batch_id else f"REQ-{r.id}"
                if b_id not in supply_batches:
                    supply_batches[b_id] = {
                        'amount': Decimal('0.00'),
                        'date': r.updated_at.date().isoformat(),
                        'action_type': r.admin_action_type or 'Supply Order'
                    }
                item_price = r.related_supply_item.unit_price or Decimal('0.00')
                subtotal = item_price * (r.requested_quantity or 1)
                supply_batches[b_id]['amount'] += subtotal
                total_asset_deduction += subtotal
            else:
                asset_deduction_details.append({
                    'id': str(r.id),
                    'item_name': "Other Asset",
                    'amount': float(amt),
                    'date': r.updated_at.date().isoformat(),
                    'action_type': r.admin_action_type or 'Asset Deduction'
                })
                total_asset_deduction += amt

        for b_id, data in supply_batches.items():
            display_name = f"Supply Asset Order ({b_id})" if b_id.startswith('BATCH') else "Supply Asset Order"
            asset_deduction_details.append({
                'id': b_id,
                'item_name': display_name,
                'amount': float(data['amount']),
                'date': data['date'],
                'action_type': data['action_type']
            })

        return Response({
            'total_days': (to_dt - from_dt).days + 1,
            'expected_working_days': m['expected_working_days'],
            'present_days': m['present_days'],
            'half_days': m['half_days'],
            'checked_in_days': m['checked_in_days'],
            'absent_days': m['absent_days'],
            'overtime_hours': m['overtime_hours'],
            'paid_leaves': m['paid_leaves'],
            'unpaid_leaves': m['unpaid_leaves'],
            'total_leaves': m['paid_leaves'] + m['unpaid_leaves'],
            'leave_details': leave_details,
            'total_reimbursement': float(total_reimbursement),
            'reimbursement_details': reimbursement_details,
            'total_asset_deduction': float(total_asset_deduction),
            'asset_deduction_details': asset_deduction_details,
        })

class SalaryDisbursementStatementView(APIView):
    """Generate a salary disbursement statement for all employees in a date range."""
    permission_classes = [IsAuthenticated, IsAdminUser | IsMaster]

    def get(self, request):
        print(f"SalaryDisbursementStatementView reached! download_excel: {request.query_params.get('download_excel')}")
        try:
            company = request.user.company
            if not company:
                return Response({'error': 'No company associated with your account.'}, status=400)

            from_date = request.query_params.get('from_date')
            to_date = request.query_params.get('to_date')

            if not from_date or not to_date:
                return Response({'error': 'from_date and to_date are required'}, status=400)

            try:
                from datetime import datetime as dt, timedelta
                from_dt = dt.strptime(from_date, '%Y-%m-%d').date()
                to_dt = dt.strptime(to_date, '%Y-%m-%d').date()
            except ValueError:
                return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

            salary_structure = SalaryStructure.objects.filter(company=company).order_by('-created_at').first()
            # Exclude employees who are also admins
            employees = Employee.objects.filter(company=company, is_active=True).exclude(user__role='admin').select_related('department', 'designation', 'shift_assigned')
            
            # Fetch Payroll Components (Earning/Deduction)
            gross_components = list(GrossSalaryComponent.objects.filter(company=company, is_active=True))
            deduction_components = list(SalaryDeductionComponent.objects.filter(company=company, is_active=True))

            # Fetch Holidays
            holiday_dates = set(CalendarEvent.objects.filter(
                company=company, is_holiday=True, date__gte=from_dt, date__lte=to_dt
            ).values_list('date', flat=True))

            # Fetch Department Working Days Config
            dept_configs = {
                cfg.department_id: cfg for cfg in DepartmentWiseWorkingDays.objects.filter(company=company)
            }

            # Pre-fetch all attendance records for the range
            attendances_qs = Attendance.objects.filter(
                employee__company=company,
                date__gte=from_dt,
                date__lte=to_dt
            )

            # Pre-fetch all approved leaves
            leaves_qs = EmpLeave.objects.filter(
                employee__company=company,
                status='Approved',
                from_date__lte=to_dt,
                to_date__gte=from_dt
            ).select_related('leave_type')

            # Pre-fetch all approved reimbursements
            reimbursements_qs = ReimbursementRequest.objects.filter(
                employee__company=company,
                status='approved',
                created_at__date__gte=from_dt,
                created_at__date__lte=to_dt
            )

            # Organize data by employee
            emp_data = []
            total_credit_amount = Decimal(0)
            
            # Pre-fetch finalized salaries for this period
            finalized_salaries = {
                fs.employee_id: fs for fs in FinalizedSalary.objects.filter(
                    company=company, from_date=from_date, to_date=to_date
                )
            }

            for emp in employees:
                # Check if we have a finalized/saved payroll for this employee in this period
                finalized = finalized_salaries.get(emp.id)
                
                if finalized:
                    # Use saved configuration (exclusions) but re-calculate using current logic
                    # to pick up any global changes in salary structure or component values.
                    saved_g_chk = finalized.config.get('gChk', {})
                    saved_d_chk = finalized.config.get('dChk', {})
                    saved_ot_enabled = finalized.config.get('otEnabled', False)
                    
                    # --- 1. Calculate Expected Working Days ---
                    cfg = dept_configs.get(emp.department_id)
                    weekend_days = [d.lower() for d in cfg.weekend_days] if cfg and cfg.weekend_days else ["saturday", "sunday"]
                    
                    expected_days = 0
                    curr = from_dt
                    while curr <= to_dt:
                        if curr.strftime('%A').lower() not in weekend_days and curr not in holiday_dates:
                            expected_days += 1
                        curr += timedelta(days=1)
                    
                    if expected_days == 0: expected_days = 30

                    # --- 2. Calculate Payable Days ---
                    emp_atts = [a for a in attendances_qs if a.employee_id == emp.id]
                    present_days, half_days, ot_hours = 0, 0, 0
                    for att in emp_atts:
                        is_holiday_work = att.date in holiday_dates or att.date.strftime('%A').lower() in weekend_days
                        w_hours = att.total_work_duration.total_seconds() / 3600 if att.total_work_duration else 0
                        
                        if is_holiday_work:
                            ot_hours += w_hours # Entire work on holiday is OT
                        else:
                            s_full = emp.shift_assigned.full_day_hours() if emp.shift_assigned else 8.0
                            s_half = emp.shift_assigned.half_day_hours() if emp.shift_assigned else 4.0
                            if w_hours >= s_full: present_days += 1
                            elif w_hours >= s_half: half_days += 1
                            if att.overtime_duration:
                                ot_hours += att.overtime_duration.total_seconds() / 3600
                    ot_hours = round(ot_hours, 2)
                    
                    emp_leaves = [l for l in leaves_qs if l.employee_id == emp.id]
                    paid_leave_days = 0
                    for lv in emp_leaves:
                        l_start, l_end = max(lv.from_date, from_dt), min(lv.to_date, to_dt)
                        d = l_start
                        while d <= l_end:
                            if lv.leave_type and lv.leave_type.is_paid: paid_leave_days += 1
                            d += timedelta(days=1)
                    
                    payable_days = Decimal(present_days) + (Decimal(half_days) * Decimal('0.5')) + Decimal(paid_leave_days)
                    
                    # --- 3. Salary Calculations ---
                    fixed_basic = emp.basic_salary or (emp.gross_salary * (salary_structure.basic_percent / Decimal(100)) if salary_structure and salary_structure.basic_percent else Decimal(0))
                    fixed_gross = emp.gross_salary or Decimal(0)
                    earned_basic = ((fixed_basic / Decimal(expected_days)) * payable_days).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    
                    hourly_rate = (fixed_basic / Decimal(expected_days)) / Decimal(8)
                    ot_pay = (hourly_rate * Decimal(ot_hours)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) if saved_ot_enabled else Decimal(0)
                    
                    # --- 4. Earnings ---
                    current_gross = earned_basic + ot_pay
                    for gc in gross_components:
                        if not saved_g_chk.get(f'g-{gc.id}', True): continue
                        amount = (earned_basic * gc.value) / Decimal(100) if gc.calc_type == 'percentage' else gc.value
                        current_gross += amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    
                    emp_reim = [r for r in reimbursements_qs if r.employee_id == emp.id]
                    current_gross += sum(Decimal(str(r.amount)) for r in emp_reim)
                    
                    # --- 5. Deductions ---
                    total_ded = Decimal(0)
                    for dc in deduction_components:
                        if not saved_d_chk.get(f'd-{dc.id}', True): continue
                        t_base = fixed_basic if dc.threshold_on == 'basic' else fixed_gross
                        if dc.has_threshold and t_base < dc.threshold_amount: continue
                        d_base = earned_basic if dc.deduct_from == 'basic' else current_gross
                        amount = (d_base * dc.value) / Decimal(100) if dc.calc_type == 'percentage' else dc.value
                        total_ded += amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    
                    net_pay = (current_gross - total_ded).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    
                    # Loan EMI deduction (Sum of all active loans)
                    # Loan EMI deduction (Period-aware)
                    loan_emi_dec = Decimal(0)
                    loans_list = LoanApplication.objects.filter(employee=emp.user, status='APPROVED')
                    for l_obj in loans_list:
                        s_date = l_obj.created_at.date()
                        m_cnt = s_date.month + l_obj.repayment_months
                        e_yr = s_date.year + (m_cnt - 1) // 12
                        e_mn = (m_cnt - 1) % 12 + 1
                        import calendar
                        _, e_lst = calendar.monthrange(e_yr, e_mn)
                        e_date = date(e_yr, e_mn, e_lst)
                        if from_dt > e_date: continue
                        
                        rep_day = s_date.day
                        c_date = from_dt
                        is_match = False
                        while c_date <= to_dt:
                            if c_date.day == rep_day:
                                is_match = True
                                break
                            _, l_d_m = calendar.monthrange(c_date.year, c_date.month)
                            if rep_day > l_d_m and c_date.day == l_d_m:
                                is_match = True
                                break
                            c_date += timedelta(days=1)
                        if is_match:
                            loan_emi_dec += Decimal(str(l_obj.emi_amount))
                    
                    net_pay = (net_pay - loan_emi_dec).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    # Loan Disbursement (Principal amount if approved in this period)
                    loan_disbursement = LoanApplication.objects.filter(
                        employee=emp.user,
                        status__in=['APPROVED', 'CLEARED'],
                        created_at__date__gte=from_dt,
                        created_at__date__lte=to_dt
                    ).aggregate(total=Sum('requested_amount'))['total'] or Decimal(0)
                    loan_disb_dec = Decimal(str(loan_disbursement))
                    net_pay += loan_disb_dec
                    # Asset Deduction from finalized record
                    asset_ded_dec = Decimal(str(finalized.asset_deduction))
                    net_pay = (net_pay - asset_ded_dec).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    if net_pay < 0: net_pay = Decimal(0)
                else:
                    # Default dynamic calculation (same as above but with all components enabled)
                    # --- 1. Calculate Expected Working Days ---
                    cfg = dept_configs.get(emp.department_id)
                    weekend_days = [d.lower() for d in cfg.weekend_days] if cfg and cfg.weekend_days else ["saturday", "sunday"]
                    
                    expected_days = 0
                    curr = from_dt
                    while curr <= to_dt:
                        if curr.strftime('%A').lower() not in weekend_days and curr not in holiday_dates:
                            expected_days += 1
                        curr += timedelta(days=1)
                    
                    if expected_days == 0:
                        expected_days = 30 # fallback to avoid div by zero

                    # --- 2. Calculate Payable Days ---
                    emp_atts = [a for a in attendances_qs if a.employee_id == emp.id]
                    present_days, half_days, ot_hours = 0, 0, 0
                    for att in emp_atts:
                        is_holiday_work = att.date in holiday_dates or att.date.strftime('%A').lower() in weekend_days
                        w_hours = (att.total_work_duration.total_seconds() / 3600) if att.total_work_duration else 0
                        
                        if is_holiday_work:
                            ot_hours += w_hours
                        else:
                            s_full = emp.shift_assigned.full_day_hours() if emp.shift_assigned else 8.0
                            s_half = emp.shift_assigned.half_day_hours() if emp.shift_assigned else 4.0
                            if w_hours >= s_full: present_days += 1
                            elif w_hours >= s_half: half_days += 1
                            if att.overtime_duration:
                                ot_hours += att.overtime_duration.total_seconds() / 3600
                    ot_hours = round(ot_hours, 2)
                    
                    emp_leaves = [l for l in leaves_qs if l.employee_id == emp.id]
                    paid_leave_days = 0
                    for lv in emp_leaves:
                        l_start = max(lv.from_date, from_dt)
                        l_end = min(lv.to_date, to_dt)
                        d = l_start
                        while d <= l_end:
                            if lv.leave_type and lv.leave_type.is_paid:
                                paid_leave_days += 1
                            d += timedelta(days=1)
                    
                    payable_days = Decimal(present_days) + (Decimal(half_days) * Decimal('0.5')) + Decimal(paid_leave_days)
                    
                    # --- 3. Salary Calculations ---
                    fixed_basic = emp.basic_salary or (emp.gross_salary * (salary_structure.basic_percent / Decimal(100)) if salary_structure and salary_structure.basic_percent else Decimal(0))
                    fixed_gross = emp.gross_salary or Decimal(0)

                    # Earned Basic (Prorated)
                    earned_basic = ((fixed_basic / Decimal(expected_days)) * payable_days).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    
                    # Overtime Pay
                    hourly_rate = (fixed_basic / Decimal(expected_days)) / Decimal(8)
                    ot_pay = (hourly_rate * Decimal(ot_hours)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    
                    # --- 4. Earnings (Gross Components) ---
                    current_gross = earned_basic + ot_pay
                    for gc in gross_components:
                        amount = (earned_basic * gc.value) / Decimal(100) if gc.calc_type == 'percentage' else gc.value
                        amount = amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                        current_gross += amount
                    
                    # Reimbursements
                    emp_reim = [r for r in reimbursements_qs if r.employee_id == emp.id]
                    current_gross += sum(Decimal(str(r.amount)) for r in emp_reim) # Reimbursements are usually fixed amounts
                    
                    # --- 5. Deductions ---
                    total_ded = Decimal(0)
                    for dc in deduction_components:
                        # Threshold check
                        t_base = fixed_basic if dc.threshold_on == 'basic' else fixed_gross
                        if dc.has_threshold and t_base < dc.threshold_amount:
                            continue
                        
                        d_base = earned_basic if dc.deduct_from == 'basic' else current_gross
                        amount = (d_base * dc.value) / Decimal(100) if dc.calc_type == 'percentage' else dc.value
                        amount = amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                        total_ded += amount
                    
                    # Final Net Pay
                    net_pay = (current_gross - total_ded).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    
                    # Loan EMI deduction (Sum of all active loans)
                    # Loan EMI deduction (Period-aware)
                    loan_emi_dec = Decimal(0)
                    loans_active = LoanApplication.objects.filter(employee=emp.user, status='APPROVED')
                    for l_act in loans_active:
                        s_dt = l_act.created_at.date()
                        m_all = s_dt.month + l_act.repayment_months
                        e_y = s_dt.year + (m_all - 1) // 12
                        e_m = (m_all - 1) % 12 + 1
                        import calendar
                        _, e_l = calendar.monthrange(e_y, e_m)
                        e_dt = date(e_y, e_m, e_l)
                        if from_dt > e_dt: continue
                        
                        r_d = s_dt.day
                        curr_v = from_dt
                        match_found = False
                        while curr_v <= to_dt:
                            _, lday_m = calendar.monthrange(curr_v.year, curr_v.month)
                            is_m = (curr_v.day == r_d)
                            if not is_m and r_d > lday_m and curr_v.day == lday_m:
                                is_m = True
                            if is_m and curr_v <= e_dt:
                                loan_emi_dec += Decimal(str(l_act.emi_amount))
                            curr_v += timedelta(days=1)
                    net_pay = (net_pay - loan_emi_dec).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    # Loan Disbursement (Principal amount if approved in this period)
                    loan_disbursement = LoanApplication.objects.filter(
                        employee=emp.user,
                        status__in=['APPROVED', 'CLEARED'],
                        created_at__date__gte=from_dt,
                        created_at__date__lte=to_dt
                    ).aggregate(total=Sum('requested_amount'))['total'] or Decimal(0)
                    loan_disb_dec = Decimal(str(loan_disbursement))
                    net_pay += loan_disb_dec
                    # Asset Deduction (Re-calc for non-finalized)
                    # Fetch approved asset requests in the range
                    asset_reqs = AssetRequest.objects.filter(
                        requested_by_id=emp.id,
                        approval_status='approved',
                        updated_at__date__range=[from_dt, to_dt]
                    )
                    
                    # Deduplicate by batch_id for supply items
                    asset_ded_dec = Decimal('0.00')
                    processed_batches = set()
                    
                    for ar in asset_reqs:
                        if ar.related_supply_item:
                            # Use (Price * Qty) for supply assets to match order total
                            item_price = ar.related_supply_item.unit_price or Decimal('0.00')
                            subtotal = item_price * (ar.requested_quantity or 1)
                            asset_ded_dec += subtotal
                        else:
                            # Core assets and others
                            asset_ded_dec += (ar.employee_payment_amount or Decimal('0'))
                            
                    net_pay = (net_pay - asset_ded_dec).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    if net_pay < 0: net_pay = Decimal(0)
                
                total_credit_amount += net_pay
                emp_data.append({
                    'id': emp.id,
                    'employee_id': emp.employee_id,
                    'full_name': emp.full_name,
                    'bank_name': emp.bank_name,
                    'account_no': emp.account_no,
                    'ifsc_code': emp.ifsc_code,
                    'gross_salary': float(current_gross.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)),
                    'loan_emi': float(loan_emi_dec),
                    'loan_disbursement': float(loan_disb_dec),
                    'asset_deduction': float(asset_ded_dec),
                    'net_salary': float(net_pay),
                    'days_paid': float(payable_days),
                })

            # Check for Excel export request
            if request.query_params.get('download_excel') == 'true':
                import openpyxl
                from openpyxl.styles import Font, Alignment, Border, Side
                from django.http import HttpResponse
                from io import BytesIO

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Salary Disbursement"

                # Header Styles
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = openpyxl.styles.PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
                alignment = Alignment(horizontal="center", vertical="center")
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                # Title and Company Info
                ws.merge_cells('A1:I1')
                ws['A1'] = f"SALARY DISBURSEMENT STATEMENT - {company.name.upper()}"
                ws['A1'].font = Font(size=14, bold=True)
                ws['A1'].alignment = Alignment(horizontal="center")
                
                ws.merge_cells('A2:I2')
                ws['A2'] = f"For the Period: {from_date} to {to_date}"
                ws['A2'].alignment = Alignment(horizontal="center")

                # Company Bank Details
                ws.merge_cells('A3:I3')
                company_bank_info = f"Company Bank: {company.bank_name or 'N/A'} | A/c No: {company.account_no or 'N/A'} | IFSC: {company.ifsc_code or 'N/A'} | Branch: {company.branch_name or 'N/A'}"
                ws['A3'] = company_bank_info
                ws['A3'].font = Font(italic=True)
                ws['A3'].alignment = Alignment(horizontal="center")

                # Table Headers
                headers = ['Sl. No', 'Employee ID', 'Employee Name', 'Bank Name', 'Account Number', 'IFSC Code', 'Gross Salary', 'Payable Days', 'Loan EMI', 'Loan Disb.', 'Asset Ded.', 'Net Salary']
                ws.append([]) # Spacer
                ws.append(headers)
                
                header_row = ws.max_row
                for cell in ws[header_row]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment
                    cell.border = border

                # Data Rows
                for idx, row_data in enumerate(emp_data, 1):
                    ws.append([
                        idx,
                        row_data['employee_id'],
                        row_data['full_name'],
                        row_data['bank_name'],
                        row_data['account_no'],
                        row_data['ifsc_code'],
                        row_data['gross_salary'],
                        row_data['days_paid'],
                        row_data['loan_emi'],
                        row_data['loan_disbursement'],
                        row_data['asset_deduction'],
                        row_data['net_salary']
                    ])
                
                last_row = ws.max_row + 1
                ws.merge_cells(f'A{last_row}:J{last_row}')
                ws[f'A{last_row}'] = "Total Disbursement Amount"
                ws[f'A{last_row}'].font = Font(bold=True)
                ws[f'A{last_row}'].alignment = Alignment(horizontal="right")
                ws[f'K{last_row}'] = float(round(total_credit_amount, 2))
                ws[f'K{last_row}'].font = Font(bold=True)

                # Column Widths
                column_widths = [8, 15, 25, 20, 20, 15, 15, 15, 15, 15, 15]
                for i, width in enumerate(column_widths):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = width

                # Return as Response
                output = BytesIO()
                wb.save(output)
                output.seek(0)

                filename = f"Salary_Disbursement_{company.name}_{from_date}.xlsx"
                response = HttpResponse(
                    output.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response

            return Response({
                'company': {
                    'name': company.name, 'address': company.address,
                    'bank_name': company.bank_name, 'account_no': company.account_no,
                    'ifsc_code': company.ifsc_code, 'branch_name': company.branch_name,
                },
                'period': { 'from_date': from_date, 'to_date': to_date },
                'summary': { 'total_employees': len(emp_data), 'total_amount': float(round(total_credit_amount, 2)) },
                'employees': emp_data
            })
        except Exception as e:
            return Response({'error': str(e)}, status=400)


class EmployeeReporteesView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        emp_id = request.data.get("employee_id")
        if not emp_id:
            return Response({"error": "employee_id is required"})
        
        try:
            manager = Employee.objects.get(id=emp_id)
        except Employee.DoesNotExist:
            return Response({"error": "Employee not found"})
        
        search = (request.query_params.get("search") or request.data.get("search") or "").strip()
        reportees = Employee.objects.filter(reporting_manager=manager)
        if search:
            reportees = reportees.filter(
                Q(first_name__icontains=search)
                | Q(middle_name__icontains=search)
                | Q(last_name__icontains=search)
                | Q(employee_id__icontains=search)
                | Q(department__department_name__icontains=search)
                | Q(designation__designation_name__icontains=search)
                | Q(status__icontains=search)
            )
        reportees = reportees.order_by("first_name", "last_name", "id")

        paginator = CustomPagination()
        paginated_qs = paginator.paginate_queryset(reportees, request, view=self)
        serializer = ReportingEmployeesSerializer(paginated_qs, many=True, context={"request": request})
        return paginator.get_paginated_response(serializer.data)


class OrganizationHierarchyView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        company = user.company
        if not company:
            return Response([])

        # Fetch all employees in the company (excluding 'master' role)
        employees = Employee.objects.filter(company=company).exclude(user__role='master').select_related('designation', 'user', 'reporting_manager')
        
        # Build map of manager_id -> list of direct reportees
        manager_map = {}
        for emp in employees:
            mid = emp.reporting_manager_id
            if mid not in manager_map:
                manager_map[mid] = []
            manager_map[mid].append(emp)

        def build_node(emp):
            return {
                'id': emp.id,
                'user_id': emp.user.id if emp.user else None,
                'name': f"{(emp.first_name or '').strip()} {(emp.last_name or '').strip()}".strip() or (emp.user.username if emp.user else "Unknown"),
                'designation': emp.designation.designation_name if emp.designation else (emp.user.role.capitalize() if emp.user else "Employee"),
                'photo': request.build_absolute_uri(emp.photo.url) if emp.photo and hasattr(emp.photo, 'url') else None,
                'email': emp.email,
                'mobile': emp.mobile,
                'employee_id': emp.employee_id,
                'department': emp.department.department_name if emp.department else None,
                'children': [build_node(child) for child in manager_map.get(emp.id, [])]
            }

        # Identify roots (those with no reporting manager or manager not in the same dataset)
        all_ids = set(emp.id for emp in employees)
        roots = [emp for emp in employees if emp.reporting_manager_id is None or emp.reporting_manager_id not in all_ids]
        
        tree = [build_node(root) for root in roots]
        return Response(tree)


class PersonalReportingLineView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        try:
            # Get the employee profile for the current user
            employee = Employee.objects.select_related('reporting_manager', 'user', 'designation').get(user=user)
        except Employee.DoesNotExist:
            return Response([])

        def build_node_simple(emp, depth=0, max_depth=2):
            """
            Builds a node and potentially its children up to a certain depth.
            depth 0: Root (Manager or Employee if no manager)
            depth 1: Children (Employee/Peers or Employee's Reportees)
            depth 2: Grandchildren (Employee's Reportees if depth 0 was Manager)
            """
            node = {
                'id': emp.id,
                'user_id': emp.user.id if emp.user else None,
                'name': f"{(emp.first_name or '').strip()} {(emp.last_name or '').strip()}".strip() or (emp.user.username if emp.user else "Unknown"),
                'designation': emp.designation.designation_name if emp.designation else (emp.user.role.capitalize() if emp.user else "Employee"),
                'photo': request.build_absolute_uri(emp.photo.url) if emp.photo and hasattr(emp.photo, 'url') else None,
                'email': emp.email,
                'mobile': emp.mobile,
                'employee_id': emp.employee_id,
                'department': emp.department.department_name if emp.department else None,
                'children': []
            }

            # Decide whether to fetch children based on depth and identity
            should_fetch_children = False
            if depth < max_depth:
                if depth == 0:
                    should_fetch_children = True # Manager always gets their reportees
                elif depth == 1 and emp.id == employee.id:
                    should_fetch_children = True # Only the current employee gets their reportees at this level

            if should_fetch_children:
                children_qs = Employee.objects.filter(reporting_manager=emp).exclude(user__role='master').select_related('user', 'designation')
                node['children'] = [build_node_simple(child, depth=depth+1, max_depth=max_depth) for child in children_qs]
            
            return node

        manager = employee.reporting_manager
        if manager:
            # Start from manager to show peers
            tree = [build_node_simple(manager, depth=0, max_depth=2)]
        else:
            # No manager, start from employee
            tree = [build_node_simple(employee, depth=0, max_depth=1)]

        return Response(tree)


# Office Structure ViewSets
class OfficeLocationViewSet(viewsets.ModelViewSet):
    serializer_class = OfficeLocationSerializer
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            permission_classes = [IsAuthenticated]
        else:
            permission_classes = [IsAuthenticated, IsAdminUser]
        return [permission() for permission in permission_classes]
    
    def get_queryset(self):
        return OfficeLocation.objects.filter(company=self.request.user.company)

    def get_serializer_context(self):
        return {'request': self.request}


class OfficeFloorViewSet(viewsets.ModelViewSet):
    serializer_class = OfficeFloorSerializer
    queryset = OfficeFloor.objects.none()  # Base queryset to satisfy DRF requirements
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            permission_classes = [IsAuthenticated]
        else:
            permission_classes = [IsAuthenticated, IsAdminUser]
        return [permission() for permission in permission_classes]
    
    def get_queryset(self):
        location_id = self.request.query_params.get('location')
        queryset = OfficeFloor.objects.filter(company=self.request.user.company).prefetch_related('sections__seats__employee')
        if location_id:
            queryset = queryset.filter(location_id=location_id)
        return queryset

    def get_serializer_context(self):
        return {'request': self.request}

    def perform_create(self, serializer):
        instance = serializer.save(company=self.request.user.company)
        self.sync_seats(instance)

    def perform_update(self, serializer):
        instance = serializer.save()
        if 'layout_data' in serializer.validated_data:
            self.sync_seats(instance)

    def sync_seats(self, floor):
        """
        Synchronizes visual seat elements from layout_data with OfficeSeat database records.
        """
        layout_data = floor.layout_data
        if not layout_data or 'elements' not in layout_data:
            return

        # 1. Get or create a default section for the floor
        # Use first available section or create a default one
        section = OfficeSection.objects.filter(floor=floor).first()
        if not section:
            section = OfficeSection.objects.create(
                floor=floor,
                name="Main Section",
                position_x=0, 
                position_y=0
            )

        # 2. Extract seat elements from layout
        layout_seats = [e for e in layout_data['elements'] if e.get('type') == 'seat']
        seat_numbers = [str(s.get('name')) for s in layout_seats if s.get('name')]

        # 3. Create or update seats found in layout
        for ls in layout_seats:
            seat_num = str(ls.get('name'))
            if not seat_num:
                continue
            
            OfficeSeat.objects.update_or_create(
                section=section,
                seat_number=seat_num,
                defaults={
                    'position_x': ls.get('x', 0),
                    'position_y': ls.get('y', 0),
                    'rotation': ls.get('rotation', 0),
                    'is_available': True
                }
            )

        # 4. Remove seats that no longer exist in the layout
        OfficeSeat.objects.filter(section__floor=floor).exclude(seat_number__in=seat_numbers).delete()


class OfficeSectionViewSet(viewsets.ModelViewSet):
    serializer_class = OfficeSectionSerializer
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            permission_classes = [IsAuthenticated]
        else:
            permission_classes = [IsAuthenticated, IsAdminUser]
        return [permission() for permission in permission_classes]
    
    def get_queryset(self):
        floor_id = self.request.query_params.get('floor')
        queryset = OfficeSection.objects.filter(floor__company=self.request.user.company).select_related('department', 'floor').prefetch_related('seats__employee')
        if floor_id:
            queryset = queryset.filter(floor_id=floor_id)
        return queryset


class OfficeSeatViewSet(viewsets.ModelViewSet):
    serializer_class = OfficeSeatSerializer
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            permission_classes = [IsAuthenticated]
        else:
            permission_classes = [IsAuthenticated, IsAdminUser]
        return [permission() for permission in permission_classes]
    
    def get_queryset(self):
        section_id = self.request.query_params.get('section')
        floor_id = self.request.query_params.get('floor')
        queryset = OfficeSeat.objects.filter(section__floor__company=self.request.user.company).select_related('employee', 'section')
        if section_id:
            queryset = queryset.filter(section_id=section_id)
        if floor_id:
            queryset = queryset.filter(section__floor_id=floor_id)
        return queryset


class SeatBookingViewSet(viewsets.ModelViewSet):
    queryset = SeatBooking.objects.all()
    serializer_class = SeatBookingSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        # For history view, we want to see even cancelled/rejected (inactive) bookings
        is_history = self.request.query_params.get('history', 'false').lower() == 'true'
        
        if is_history:
            queryset = SeatBooking.objects.all()
        else:
            queryset = SeatBooking.objects.filter(is_active=True)
        
        # Enforce company security: Admins and Employees should only see bookings from their own company
        if self.request.user.role in ['admin', 'employee']:
            if self.request.user.company:
                queryset = queryset.filter(employee__company=self.request.user.company)
            else:
                # If for some reason an admin/employee has no company, they see nothing
                queryset = queryset.none()
        elif self.request.user.role == 'master':
            # Masters might want to filter by company if provided
            target_company = self.request.query_params.get('company_id')
            if target_company:
                queryset = queryset.filter(employee__company_id=target_company)

        date_str = self.request.query_params.get('date', None)
        floor_id = self.request.query_params.get('floor', None)
        seat_number = self.request.query_params.get('seat_number', None)
        status = self.request.query_params.get('status', None)
        start_time_str = self.request.query_params.get('start_time', None)
        end_time_str = self.request.query_params.get('end_time', None)
        
        # When viewing the map, we usually only care about already approved bookings
        # Unless we are in history mode, in which case we show everything requested
        if (date_str or floor_id) and not is_history:
            queryset = queryset.filter(status='approved')
        
        if status:
            queryset = queryset.filter(status=status)
            
        if date_str:
            try:
                target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                queryset = queryset.filter(
                    Q(start_date__lte=target_date),
                    Q(end_date__gte=target_date) | Q(end_date__isnull=True)
                )
            except ValueError:
                pass

        if start_time_str and end_time_str:
            try:
                target_start = datetime.strptime(start_time_str, '%H:%M').time()
                target_end = datetime.strptime(end_time_str, '%H:%M').time()
                
                queryset = queryset.filter(
                    Q(start_time__lt=target_end) | Q(start_time__isnull=True),
                    Q(end_time__gt=target_start) | Q(end_time__isnull=True)
                )
            except ValueError:
                pass
        
        if floor_id:
            queryset = queryset.filter(seat__section__floor_id=floor_id)
            
        if seat_number:
            queryset = queryset.filter(seat__seat_number=seat_number)
            
        if not date_str and not floor_id and not status and not is_history and hasattr(self.request.user, 'employee'):
            # Show all bookings for the current employee (including pending)
            queryset = queryset.filter(employee=self.request.user.employee)
        
        return queryset.select_related('employee', 'seat__section__floor').order_by('-created_at')

    def perform_create(self, serializer):
        if not hasattr(self.request.user, 'employee'):
            raise serializers.ValidationError({"detail": "Only employees can book seats."})
            
        employee = self.request.user.employee
        booking_type = serializer.validated_data.get('booking_type', 'daily')
        start_date = serializer.validated_data.get('start_date')
        end_date = serializer.validated_data.get('end_date')

        # Logic for end_date based on type
        if booking_type == 'permanent':
            end_date = None
        elif booking_type == 'weekly' and not end_date:
            end_date = start_date + timedelta(days=6)
        elif booking_type == 'daily':
            end_date = start_date

        # Auto-approve daily bookings
        status = 'approved' if booking_type == 'daily' else 'pending'

        serializer.save(employee=employee, end_date=end_date, status=status)

    @action(detail=True, methods=['post'], permission_classes=[IsAdminUser | IsMaster])
    def approve(self, request, pk=None):
        booking = self.get_object()
        booking.status = 'approved'
        booking.save()
        return Response({'status': 'booking approved'})

    @action(detail=True, methods=['post'], permission_classes=[IsAdminUser | IsMaster])
    def reject(self, request, pk=None):
        booking = self.get_object()
        booking.status = 'rejected'
        booking.is_active = False # Rejection cancels the booking
        booking.save()
        return Response({'status': 'booking rejected'})

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def cancel(self, request, pk=None):
        booking = self.get_object()
        user = request.user
        
        # Check if user is admin/master OR the owner of the booking
        is_admin_or_master = user.role in ['admin', 'master']
        is_owner = hasattr(user, 'employee') and booking.employee == user.employee
        
        if not (is_admin_or_master or is_owner):
            return Response({'error': 'You do not have permission to cancel this booking.'}, status=status.HTTP_403_FORBIDDEN)
            
        booking.status = 'cancelled'
        booking.is_active = False
        booking.save()
        return Response({'status': 'booking cancelled'})


# --------------------------- CONFERENCE ROOM MANAGEMENT ---------------------------------

class ConferenceRoomViewSet(viewsets.ModelViewSet):
    serializer_class = ConferenceRoomSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.role == 'master':
            return ConferenceRoom.objects.all().order_by('name')
        return ConferenceRoom.objects.filter(company=user.company).order_by('name')

    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company)

class ConferenceRoomBookingViewSet(viewsets.ModelViewSet):
    serializer_class = ConferenceRoomBookingSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = ConferenceRoomBooking.objects.all()
        
        # Filtering
        status = self.request.query_params.get('status')
        date = self.request.query_params.get('date')
        is_history = self.request.query_params.get('history') == 'true'
        room_id = self.request.query_params.get('room')
        floor_id = self.request.query_params.get('floor')
        
        if user.role != 'master':
            queryset = queryset.filter(room__company=user.company)
            
        if status:
            queryset = queryset.filter(status=status)
        if date:
            queryset = queryset.filter(date=date)
        if room_id:
            queryset = queryset.filter(room_id=room_id)
        if floor_id:
            queryset = queryset.filter(room__floor_id=floor_id)
            
        if is_history:
             queryset = queryset.filter(date__lt=timezone.now().date())
        
        if not status and not date and not is_history and hasattr(user, 'employee'):
            # Show active and future bookings for the current employee
            queryset = queryset.filter(employee=user.employee, date__gte=timezone.now().date())

        return queryset.select_related('employee', 'room__floor').order_by('-date', '-start_time')

    def perform_create(self, serializer):
        user = self.request.user
        if not hasattr(user, 'employee'):
            raise serializers.ValidationError({"detail": "Only employees can book conference rooms."})
            
        employee = user.employee
        room = serializer.validated_data['room']
        start_time = serializer.validated_data['start_time']
        end_time = serializer.validated_data['end_time']
        date = serializer.validated_data['date']

        # Conflict check - Room
        if ConferenceRoomBooking.objects.filter(
            room=room,
            date=date,
            status__in=['pending', 'approved'],
            start_time__lt=end_time,
            end_time__gt=start_time
        ).exists():
            raise serializers.ValidationError({"detail": "Room is already booked for this time."})

        # Conflict check - Employee (One person cannot book multiple rooms at same time)
        if ConferenceRoomBooking.objects.filter(
            employee=employee,
            date=date,
            status__in=['pending', 'approved'],
            start_time__lt=end_time,
            end_time__gt=start_time
        ).exists():
            raise serializers.ValidationError({"detail": "You already have another room booked during this overlapping time period."})

        # Duration check for approval
        config, created = ConferenceRoomConfig.objects.get_or_create(company=user.company)
        limit = config.approval_limit_minutes
        
        # Calculate duration in minutes
        duration = (datetime.combine(date, end_time) - datetime.combine(date, start_time)).total_seconds() / 60
        
        # If duration is negative or zero
        if duration <= 0:
            raise serializers.ValidationError({"detail": "End time must be after start time."})

        status = 'approved' if duration <= limit else 'pending'
        
        serializer.save(employee=employee, status=status)

    @action(detail=True, methods=['post'], permission_classes=[IsAdminUser | IsMaster])
    def approve(self, request, pk=None):
        booking = self.get_object()
        booking.status = 'approved'
        booking.save()
        return Response({'status': 'booking approved'})

    @action(detail=True, methods=['post'], permission_classes=[IsAdminUser | IsMaster])
    def reject(self, request, pk=None):
        booking = self.get_object()
        booking.status = 'rejected'
        booking.save()
        return Response({'status': 'booking rejected'})

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def cancel(self, request, pk=None):
        booking = self.get_object()
        user = request.user
        
        # Check if user is admin/master OR the owner of the booking
        is_admin_or_master = user.role in ['admin', 'master']
        is_owner = hasattr(user, 'employee') and booking.employee == user.employee
        
        if not (is_admin_or_master or is_owner):
            return Response({'error': 'You do not have permission to cancel this booking.'}, status=status.HTTP_403_FORBIDDEN)
            
        booking.status = 'cancelled'
        booking.save()
        return Response({'status': 'booking cancelled'})

class ConferenceRoomConfigViewSet(viewsets.ModelViewSet):
    serializer_class = ConferenceRoomConfigSerializer
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated(), (IsAdminUser | IsMaster)()]

    def get_queryset(self):
        return ConferenceRoomConfig.objects.filter(company=self.request.user.company)

    def perform_create(self, serializer):
        # Ensure only one config per company
        if ConferenceRoomConfig.objects.filter(company=self.request.user.company).exists():
             raise serializers.ValidationError({"detail": "Config already exists for this company."})
        serializer.save(company=self.request.user.company)


from google.oauth2 import id_token
from google.auth.transport import requests as google_requests
from django.conf import settings
from rest_framework.permissions import AllowAny

def _normalize_domain_list(raw: str | None) -> list[str]:
    if not raw:
        return []
    parts = []
    for x in str(raw).split(","):
        d = (x or "").strip().lower()
        if d.startswith("@"):
            d = d[1:]
        if d:
            parts.append(d)
    return parts


def _company_for_email_domain(email: str):
    domain = (email.split("@")[-1] if "@" in email else "").strip().lower()
    if not domain:
        return None
    # gmail_domains is comma-separated; easiest safe match is to scan companies.
    for c in Company.objects.exclude(gmail_domains__isnull=True).exclude(gmail_domains__exact="").only("id", "gmail_domains"):
        if domain in _normalize_domain_list(c.gmail_domains):
            return c
    return None


def _ensure_employee_for_company_user(*, user: UserRegister, email: str, first_name: str = "", last_name: str = ""):
    """
    If user logs in via SSO for the first time and their email domain matches a company's allowed domains,
    auto-create/link an Employee profile under that company.
    """
    if user.role != "employee":
        return
    if getattr(user, "company_id", None):
        return

    company = _company_for_email_domain(email)
    if not company:
        return

    # Link user to company
    user.company_id = company.id
    user.save(update_fields=["company"])

    # Link or create Employee
    emp = Employee.objects.filter(email__iexact=email).select_related("company").first()
    if emp:
        if not emp.company_id:
            emp.company_id = company.id
        if not emp.user_id:
            emp.user_id = user.id
        if not emp.first_name and first_name:
            emp.first_name = first_name
        if not emp.last_name and last_name:
            emp.last_name = last_name
        emp.save()
        return

    Employee.objects.create(
        user_id=user.id,
        company_id=company.id,
        email=email,
        first_name=first_name or None,
        last_name=last_name or None,
    )


class GoogleLoginAPIView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        token = request.data.get("credential")
        if not token:
            return Response({"detail": "Credential missing"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Verify Google Token
            idinfo = id_token.verify_oauth2_token(
                token, 
                google_requests.Request(), 
                settings.GOOGLE_CLIENT_ID
            )
            
            email = (idinfo.get("email") or "").strip().lower()
            first_name = idinfo.get("given_name", "")
            last_name = idinfo.get("family_name", "")
            
            # Determine role for new user:
            # If NO users exist on the platform yet, the first SSO user becomes 'master'.
            # All subsequent users default to 'employee'.
            if UserRegister.objects.count() == 0:
                default_role = "master"
            else:
                default_role = "employee"

            # Find or Create User
            user, created = UserRegister.objects.get_or_create(
                email=email,
                defaults={
                    "username": email,
                    "first_name": first_name,
                    "last_name": last_name,
                    "role": default_role,
                    "is_active": True
                }
            )
            if created:
                user.set_unusable_password()
                user.save()


            # If this is an employee login, ensure profile is synced and ID generated
            if user.role == "employee":
                emp = Employee.objects.filter(email__iexact=email).first()
                if emp:
                    # Link to User account if needed
                    if not emp.user_id:
                        emp.user_id = user.id
                    # Link User to Company if missing
                    if not getattr(user, "company_id", None) and emp.company_id:
                        user.company_id = emp.company_id
                        user.save(update_fields=["company"])
                    # Trigger model save (generates employee_id if missing)
                    emp.save()
                elif not getattr(user, "company_id", None):
                    # If no Employee profile yet, try auto-link via company domain
                    _ensure_employee_for_company_user(user=user, email=email, first_name=first_name, last_name=last_name)

            # Return tokens
            refresh = RefreshToken.for_user(user)
            return Response({
                "access": str(refresh.access_token),
                "refresh": str(refresh),
                "id": user.id,
                "username": user.username,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "role": user.role,
                "is_reporting_manager": user.is_reporting_manager
            }, status=status.HTTP_200_OK)

        except ValueError as e:
            print("Google verify error:", str(e))
            return Response({"detail": f"Invalid Google token: {str(e)}"}, status=status.HTTP_401_UNAUTHORIZED)
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class LoanCategoryViewSet(viewsets.ModelViewSet):
    serializer_class = LoanCategorySerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination

    def get_queryset(self):
        user = self.request.user
        if user.role == 'master':
            return LoanCategory.objects.all().order_by('-created_at')
        if hasattr(user, 'company') and user.company:
            return LoanCategory.objects.filter(company=user.company).order_by('-created_at')
        return LoanCategory.objects.none()

    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company)

class LoanApplicationViewSet(viewsets.ModelViewSet):
    queryset = LoanApplication.objects.all()
    serializer_class = LoanApplicationSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['category__name', 'status', 'employee__first_name', 'employee__last_name']

    def destroy(self, request, *args, **kwargs):
        user = request.user
        instance = self.get_object()
        if user.role in ['admin', 'master']:
            return super().destroy(request, *args, **kwargs)
        if instance.employee == user and instance.status == 'PENDING':
            return super().destroy(request, *args, **kwargs)
        from rest_framework.response import Response
        return Response({"error": "You do not have permission to delete this loan."}, status=403)

    def get_queryset(self):
        user = self.request.user
        qs = LoanApplication.objects.filter(employee__company=user.company).order_by('-created_at')
        
        if user.role in ['admin', 'master']:
            # Admins see:
            # 1. All non-PENDING applications (APPROVED, REJECTED, MANAGER_APPROVED)
            # 2. PENDING applications ONLY IF the applicant's reporting manager is NOT an employee
            # (meaning they report to an admin or have no manager assigned)
            return qs.exclude(
                status='PENDING', 
                employee__employee__reporting_manager__user__role='employee'
            )
            
        # For managers (even if they are employees), show their reportees' applications + their own
        if user.is_reporting_manager:
            try:
                employee_profile = user.employee_profile
                reportee_user_ids = Employee.objects.filter(reporting_manager=employee_profile).values_list('user_id', flat=True)
                # Combine reportees' applications and user's own applications
                return qs.filter(employee_id__in=reportee_user_ids) | qs.filter(employee=user)
            except:
                return qs.filter(employee=user)
        
        # Regular employees only see their own
        return qs.filter(employee=user)

    @action(detail=False, methods=['post'])
    def check_eligibility(self, request):
        employee = request.user.employee_profile
        category_id = request.data.get('category_id')
        amount = float(request.data.get('amount', 0))
        
        try:
            category = LoanCategory.objects.get(id=category_id, company=request.user.company)
        except LoanCategory.DoesNotExist:
            return Response({"eligible": False, "reason": "Invalid loan category"}, status=400)

        # 0. Active Loan Check
        active_loan = LoanApplication.objects.filter(
            employee=request.user, 
            status__in=['PENDING', 'MANAGER_APPROVED', 'APPROVED']
        ).exists()
        if active_loan:
            return Response({
                "eligible": False, 
                "reason": "You already have an active or pending loan. Please clear your current loan before applying for a new one."
            })

        # 1. Tenure Check
        if employee.date_of_joining:
            from datetime import date
            tenure_months = (date.today() - employee.date_of_joining).days // 30
            if tenure_months < category.min_tenure_months:
                return Response({
                    "eligible": False, 
                    "reason": f"Minimum company tenure required: {category.min_tenure_months} months. Your tenure: {tenure_months} months."
                })
        else:
            return Response({"eligible": False, "reason": "Date of joining not set. Please contact HR."})

        # 2. Level Check
        if not category.allowed_levels.filter(id=employee.level_id).exists():
            return Response({"eligible": False, "reason": f"Your level ({employee.level.level_name}) is not eligible for this loan type."})

        # 3. Max Amount Check
        if amount > float(category.max_loan_limit):
            return Response({"eligible": False, "reason": f"Requested amount exceeds the maximum limit of ₹{category.max_loan_limit} for this loan."})

        # 4. Interest Slab
        slab = category.interest_slabs.filter(min_amount__lte=amount, max_amount__gte=amount).first()
        if not slab:
            # Fallback to highest slab if amount exceeds all slabs but is within max limit
            slab = category.interest_slabs.order_by('-max_amount').first()
        
        rate = float(slab.interest_rate) if slab else 0

        return Response({
            "eligible": True,
            "interest_rate": rate,
            "max_repayment_months": category.max_repayment_months
        })

    def perform_create(self, serializer):
        # Final safety check for active loans
        active_loan = LoanApplication.objects.filter(
            employee=self.request.user, 
            status__in=['PENDING', 'MANAGER_APPROVED', 'APPROVED']
        ).exists()
        if active_loan:
            from rest_framework.exceptions import ValidationError
            raise ValidationError("You already have an active loan. Please clear it before applying for a new one.")

        # Additional validation on save
        employee = self.request.user.employee_profile
        category = serializer.validated_data['category']
        amount = serializer.validated_data['requested_amount']
        months = serializer.validated_data['repayment_months']
        
        # Calculate EMI and Rate
        slab = category.interest_slabs.filter(min_amount__lte=amount, max_amount__gte=amount).first()
        if not slab:
            slab = category.interest_slabs.order_by('-max_amount').first()
        
        rate = slab.interest_rate if slab else 0
        
        # Simple EMI calculation: (Principal + Total Interest) / Months
        total_interest = (float(amount) * float(rate) * (months / 12)) / 100
        emi = (float(amount) + total_interest) / months
        
        serializer.save(
            employee=self.request.user,
            interest_rate=rate,
            emi_amount=emi
        )

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        application = self.get_object()
        user = request.user
        status = request.data.get('status') # 'APPROVED' or 'REJECTED'
        remarks = request.data.get('remarks', '')

        if status == 'REJECTED':
            application.status = 'REJECTED'
            application.admin_remarks = remarks
            application.save()
            return Response({"message": "Application rejected"})

        # Approval Logic
        is_admin = user.role == 'admin'
        
        # Check if user is the reporting manager
        is_manager = False
        try:
            manager_emp = user.employee_profile
            if application.employee.employee_profile.reporting_manager == manager_emp:
                is_manager = True
        except:
            pass

        if is_manager and application.status == 'PENDING':
            # Manager Approval
            application.manager_approved_by = user
            if is_admin:
                # If manager is also admin, approve fully
                application.status = 'APPROVED'
                application.admin_approved_by = user
            else:
                application.status = 'MANAGER_APPROVED'
            application.save()
            return Response({"message": "Manager approved. Pending Admin approval." if application.status == 'MANAGER_APPROVED' else "Loan fully approved."})

        if is_admin and application.status in ['MANAGER_APPROVED', 'PENDING']:
            # Admin Approval
            application.admin_approved_by = user
            application.status = 'APPROVED'
            application.admin_remarks = remarks
            application.save()
            return Response({"message": "Loan fully approved."})

        return Response({"error": "You do not have permission to approve this application at this stage."}, status=403)
        
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        user = request.user
        application = self.get_object()
        remarks = request.data.get('remarks', '')
        
        is_admin = user.role in ['admin', 'master']
        is_manager = False
        try:
            manager_emp = user.employee_profile
            if application.employee.employee_profile.reporting_manager == manager_emp:
                is_manager = True
        except:
            pass

        if is_admin or is_manager:
            application.status = 'REJECTED'
            application.admin_remarks = remarks
            application.save()
            return Response({"status": "REJECTED"})
            
        return Response({"error": "Not authorized"}, status=403)

    @action(detail=True, methods=['post'])
    def clear(self, request, pk=None):
        """Mark an approved loan as cleared (repaid)."""
        if request.user.role not in ['admin', 'master']:
            return Response({"error": "Only admins can clear loans"}, status=403)
            
        application = self.get_object()
        if application.status != 'APPROVED':
            return Response({"error": "Only approved loans can be cleared"}, status=400)
            
        application.status = 'CLEARED'
        application.save()
        return Response({"status": "CLEARED"})

class LoanInterestSlabViewSet(viewsets.ModelViewSet):
    serializer_class = LoanInterestSlabSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        category_id = self.request.query_params.get('category_id')
        if category_id:
            return LoanInterestSlab.objects.filter(category_id=category_id)
        return LoanInterestSlab.objects.all()

class WFHRequestViewSet(viewsets.ModelViewSet):
    serializer_class = WFHRequestSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination

    def get_queryset(self):
        user = self.request.user
        mine = self.request.query_params.get('mine') == 'true'
        
        emp_profile = user.employee_profile
        if not emp_profile:
            # If a user doesn't have an employee profile, they can't request or approve WFH
            # (Admins/Masters usually have a linked profile if they manage people)
            return WFHRequest.objects.none()

        if mine:
            # ONLY show requests where user is the requester
            return WFHRequest.objects.filter(employee=emp_profile).order_by('-created_at')
            
        # APPROVALS: Only show requests where current user is the reporting manager
        return WFHRequest.objects.filter(reporting_manager=emp_profile).order_by('-created_at')

    def perform_create(self, serializer):
        emp_profile = self.request.user.employee_profile
        if not emp_profile:
            raise serializers.ValidationError("User must have an employee profile to request WFH.")
        
        # Automatically assign the employee's reporting manager
        serializer.save(
            employee=emp_profile,
            reporting_manager=emp_profile.reporting_manager
        )

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        wfh_request = self.get_object()
        
        # Permission check: Only reporting manager or admin can approve
        if request.user.role not in ['admin', 'master'] and request.user.employee_profile != wfh_request.reporting_manager:
            return Response({"detail": "You do not have permission to approve this request."}, status=status.HTTP_403_FORBIDDEN)

        if wfh_request.status != 'pending':
            return Response({"detail": "Only pending requests can be approved."}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            wfh_request.status = 'approved'
            wfh_request.save()

            employee = wfh_request.employee
            old_loc = employee.work_location
            new_loc = 'office' if wfh_request.request_type == 'wfo' else 'home'

            employee.work_location = new_loc
            employee.save()

            req_type_str = "Work From Office" if wfh_request.request_type == 'wfo' else "Work From Home"

            # Create log
            WorkLocationLog.objects.create(
                employee=employee,
                from_location=old_loc,
                to_location=new_loc,
                changed_by=request.user.employee_profile,
                reason=f"{req_type_str} Request Approved: {wfh_request.reason}"
            )

        return Response({"status": "approved", "work_location": new_loc})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        wfh_request = self.get_object()
        rejection_reason = request.data.get('rejection_reason', 'No reason provided')

        # Permission check
        if request.user.role not in ['admin', 'master'] and request.user.employee_profile != wfh_request.reporting_manager:
            return Response({"detail": "You do not have permission to reject this request."}, status=status.HTTP_403_FORBIDDEN)

        wfh_request.status = 'rejected'
        wfh_request.rejection_reason = rejection_reason
        wfh_request.save()

        return Response({"status": "rejected"})

class WorkLocationLogViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = WorkLocationLogSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return WorkLocationLog.objects.all().order_by('-date')


